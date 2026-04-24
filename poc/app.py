import json
import os
import re
import uuid
import hashlib
import secrets as _secrets
import threading
from datetime import datetime, timezone
from flask import Flask, render_template, request, jsonify, Response, stream_with_context, send_file
import anthropic
import rag
import cabinet_sizing as cs
import catalogue_matcher as cm
import file_parser as fp
from json_repair import repair_json

# Load .env file if present
env_path = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                val = v.strip().strip('"').strip("'")
                if val:  # always overwrite with non-empty values from .env
                    os.environ[k.strip()] = val

os.environ.setdefault("FLASK_SKIP_DOTENV", "1")   # we load .env ourselves above
app = Flask(__name__)

# ── RAG index — auto-discover & build on startup ───────────────────────────────
_rag_ready      = False
_active_years: list[str] = []   # years successfully indexed (updated incrementally)
_boot_complete  = False          # True once _boot_rag finishes (all years attempted)
_boot_status    = "starting"     # human-readable current step

BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
YEARLY_DATA_DIR = os.path.join(BASE_DIR, "..", "yearly_data")

def _discover_years() -> list[str]:
    """
    Return sorted list of years available. Tries, in order:
      1. yearly_data/YYYY/ folders (raw data, for full parse)
      2. data/yearly_projects_YYYY.json files (pre-parsed)
      3. ChromaDB collections named yearly_projects_YYYY (pre-indexed)
    In a deployed environment without raw Excel files, steps 2-3 keep the app working.
    """
    years: set[str] = set()

    # 1. Raw folders
    ydir = os.path.abspath(YEARLY_DATA_DIR)
    if os.path.isdir(ydir):
        for d in os.listdir(ydir):
            if os.path.isdir(os.path.join(ydir, d)) and d.isdigit() and len(d) == 4:
                years.add(d)
        print(f"[RAG] yearly_data/ scan: {sorted(years) or '(empty)'}")
    else:
        print(f"[RAG] yearly_data/ not found (OK in deployed env): {ydir}")

    # 2. Pre-parsed JSON files
    data_dir = os.path.join(BASE_DIR, "data")
    if os.path.isdir(data_dir):
        for f in os.listdir(data_dir):
            m = re.match(r"yearly_projects_(\d{4})\.json$", f)
            if m:
                years.add(m.group(1))

    # 3. ChromaDB pre-built collections
    try:
        for y in rag.get_available_yearly_indices():
            if y.isdigit() and len(y) == 4:
                years.add(y)
    except Exception:
        pass

    result = sorted(years)
    print(f"[RAG] Total years available (raw + json + chroma): {result}")
    return result

def _check_dependencies() -> None:
    """Warn about missing optional packages at startup."""
    missing = []
    try:
        import extract_msg  # noqa
    except ImportError:
        missing.append("extract-msg  (needed to parse .msg Outlook files)")
    try:
        from docx import Document  # noqa
    except ImportError:
        missing.append("python-docx  (needed to parse .docx files)")
    if missing:
        print("\n⚠️  Missing packages — install with: pip install -r poc/requirements.txt")
        for m in missing:
            print(f"   • {m}")
        print()

def _boot_rag() -> None:
    """
    Run at startup (in background thread):
      1. Build historical-quotes index if missing.
      2. For every year folder in yearly_data/:
           a. Parse raw files → JSON if JSON missing.
           b. Build ChromaDB yearly index if not indexed.
      3. Print a startup report.
    _active_years is updated incrementally so the frontend sees progress.
    """
    global _rag_ready, _active_years, _boot_complete, _boot_status

    _check_dependencies()
    print("\n" + "═" * 55)
    print("  CETIE RAG — startup indexing")
    print("═" * 55)

    # ── Historical quotes index ──────────────────────────────
    _boot_status = "Building quotes index…"
    try:
        if not rag.is_index_ready():
            print("[RAG] Building historical-quotes index …")
            rag.build_index()
        else:
            print("[RAG] Historical-quotes index  ✓ (already built)")
        _rag_ready = True
    except Exception as e:
        print(f"[RAG] ✗ Could not build quotes index: {e}")

    # ── Yearly project indices ───────────────────────────────
    years = _discover_years()
    if not years:
        print(f"[RAG] No year folders found in {os.path.abspath(YEARLY_DATA_DIR)}")
    else:
        print(f"[RAG] Year folders detected: {', '.join(years)}")

    for year in years:
        _boot_status = f"Indexing {year}…"
        json_path = os.path.join(BASE_DIR, "data", f"yearly_projects_{year}.json")

        # Step 1: parse raw files → JSON (only if JSON missing)
        if not os.path.exists(json_path):
            print(f"[RAG] [{year}] JSON missing — parsing raw files …")
            try:
                import parse_yearly_data as pyd
                pyd.run(year=year, force=False)
            except Exception as e:
                print(f"[RAG] [{year}] ✗ Parse failed: {e}")
                continue
        else:
            print(f"[RAG] [{year}] JSON found: {json_path}")

        # Step 2: build vector index (only if not already indexed)
        if not rag.is_yearly_index_ready(year):
            print(f"[RAG] [{year}] Building vector index …")
            try:
                rag.build_yearly_index(year=year)
            except Exception as e:
                print(f"[RAG] [{year}] ✗ Index failed: {e}")
                continue
        else:
            print(f"[RAG] [{year}] Vector index already exists")

        # Verify and add immediately so frontend sees it
        if rag.is_yearly_index_ready(year):
            if year not in _active_years:
                _active_years.append(year)
                _active_years.sort()
            print(f"[RAG] [{year}] ✓ Ready  (active years so far: {_active_years})")
        else:
            print(f"[RAG] [{year}] ✗ Index not ready after build")

    # ── Done ─────────────────────────────────────────────────
    _boot_complete = True
    _boot_status   = "ready"
    print("═" * 55)
    if _active_years:
        print(f"  Active years : {', '.join(_active_years)}")
    else:
        print("  Active years : none (no yearly data indexed)")
    print(f"  Quotes index : {'✓' if _rag_ready else '✗'}")
    print(f"  yearly_data/ : {os.path.abspath(YEARLY_DATA_DIR)}")
    print("═" * 55 + "\n")

def _start_boot(force_years: list[str] | None = None):
    """Start (or restart) the RAG boot thread. force_years rebuilds those years."""
    global _boot_complete, _boot_status
    _boot_complete = False
    _boot_status   = "starting"
    t = threading.Thread(target=_boot_rag, daemon=True, name="rag-boot")
    t.start()

# Launch indexing in background so the web server starts immediately
_start_boot()


def _retrieve_yearly_projects(query_text: str, n_per_year: int = 5) -> list[dict]:
    """Query all active yearly indices and return merged, deduplicated results."""
    if not _active_years:
        return []
    results = []
    seen_ids = set()
    for year in _active_years:
        try:
            hits = rag.retrieve_similar_projects(query_text, year=year, n_results=n_per_year)
            for h in hits:
                uid = f"{year}_{h.get('id', '')}"
                if uid not in seen_ids:
                    seen_ids.add(uid)
                    h["_year"] = year
                    results.append(h)
        except Exception as e:
            print(f"[RAG] Retrieval error for year {year}: {e}")
    # Sort by similarity descending, keep top n_per_year overall
    results.sort(key=lambda x: x.get("similarity_score", 0), reverse=True)
    return results[:n_per_year]


def _ensure_rag_index():
    """No-op kept for compatibility — indexing now happens at startup."""
    pass

# Load data
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

with open(os.path.join(DATA_DIR, "blocks.json"), encoding="utf-8") as f:
    ALL_BLOCKS = json.load(f)

with open(os.path.join(DATA_DIR, "armoires.json"), encoding="utf-8") as f:
    ALL_ARMOIRES = json.load(f)

# ── Runtime state files — persistent across deploys ──────────────────────────
# In production (Render) these live on a mounted persistent disk so that
# redeploying the app does NOT wipe history, feedback, or learned rules.
# Set CETIE_STATE_DIR env var to the disk mount path (e.g. /var/cetie-state).
# Locally, falls back to DATA_DIR so dev runs work unchanged.
_STATE_DIR = os.environ.get("CETIE_STATE_DIR") or DATA_DIR
try:
    os.makedirs(_STATE_DIR, exist_ok=True)
except Exception as _e:
    print(f"[state] Could not create {_STATE_DIR}: {_e} — falling back to {DATA_DIR}")
    _STATE_DIR = DATA_DIR


def _state_path(name: str, migrate_from: str | None = None) -> str:
    """
    Return the path a runtime state file should live at. On first launch after
    switching to a persistent disk, migrate an existing file from DATA_DIR so
    you don't lose data that was baked into the repo during early testing.
    """
    persistent = os.path.join(_STATE_DIR, name)
    if migrate_from and not os.path.exists(persistent) and os.path.exists(migrate_from):
        try:
            import shutil
            shutil.copy2(migrate_from, persistent)
            print(f"[state] Migrated {migrate_from} → {persistent}")
        except Exception as e:
            print(f"[state] Could not migrate {migrate_from}: {e}")
    return persistent


FEEDBACK_PATH = _state_path("feedback.json",      migrate_from=os.path.join(DATA_DIR, "feedback.json"))
RULES_PATH    = _state_path("learned_rules.json", migrate_from=os.path.join(DATA_DIR, "learned_rules.json"))
USERS_PATH    = _state_path("users.json",         migrate_from=os.path.join(DATA_DIR, "users.json"))
HISTORY_PATH  = _state_path("history.json",       migrate_from=os.path.join(DATA_DIR, "history.json"))
print(f"[state] Runtime state directory: {_STATE_DIR}")

# Thread-safe file access (supports multiple concurrent users)
_file_lock = threading.Lock()

# ── Session store  {token: user_dict} — persisted on the state disk ─────────
# Sessions used to live only in RAM, which meant every Render redeploy kicked
# every user out (they had a valid token in their browser, but the server had
# forgotten it). We now back it with a JSON file on the persistent disk so
# tokens survive across deploys.
SESSIONS_PATH = _state_path("sessions.json")

def _load_sessions() -> dict:
    try:
        with open(SESSIONS_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _save_sessions() -> None:
    try:
        with _file_lock:
            with open(SESSIONS_PATH, "w", encoding="utf-8") as f:
                json.dump(_sessions, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[auth] Could not persist sessions: {e}")

_sessions: dict = _load_sessions()
print(f"[auth] Restored {len(_sessions)} session(s) from {SESSIONS_PATH}")

def _load_json(path, default):
    try:
        with _file_lock:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        return default

def _save_json(path, data):
    with _file_lock:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

# ── Password helpers ─────────────────────────────────────────────────────────

def _hash_password(password: str) -> str:
    salt = os.urandom(16).hex()
    dk   = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 200_000)
    return f"{salt}:{dk.hex()}"

def _verify_password(password: str, stored: str) -> bool:
    try:
        salt, dk_hex = stored.split(":", 1)
        dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 200_000)
        return dk.hex() == dk_hex
    except Exception:
        return False

def _ensure_admin():
    """Create a default admin account if the users file is empty / missing."""
    users = _load_json(USERS_PATH, [])
    if not any(u.get("role") == "admin" for u in users):
        admin = {
            "id":            "admin",
            "username":      "admin",
            "name":          "Administrator",
            "role":          "admin",
            "password_hash": _hash_password("admin123"),
            "created_at":    datetime.now(timezone.utc).isoformat(),
        }
        users.append(admin)
        _save_json(USERS_PATH, users)
        print("[Auth] Default admin created — username: admin  password: admin123")

_ensure_admin()   # runs once at startup

# ── Auth helpers ──────────────────────────────────────────────────────────────

def _current_user() -> dict | None:
    token = request.headers.get("X-Auth-Token", "")
    return _sessions.get(token)

def _require_auth():
    """Return (user, None) or (None, error_response)."""
    user = _current_user()
    if not user:
        return None, (jsonify({"error": "Not authenticated"}), 401)
    return user, None

def _require_admin():
    user, err = _require_auth()
    if err:
        return None, err
    if user.get("role") != "admin":
        return None, (jsonify({"error": "Admin required"}), 403)
    return user, None

def _safe_user(u: dict) -> dict:
    """Strip password_hash before sending to client."""
    return {k: u[k] for k in ("id","username","name","role","created_at") if k in u}

def get_learned_rules() -> str:
    """Return active learned rules formatted for prompt injection."""
    rules = _load_json(RULES_PATH, [])
    active = [r for r in rules if r.get("active", True)]
    if not active:
        return ""
    lines = "\n".join(f"{i+1}. [{r.get('scope','general').upper()}] {r['rule']}" for i, r in enumerate(active))
    return f"\nCETIE ENGINEERING RULES (learned from expert feedback — always follow strictly):\n{lines}\n"

# ── Accessory rules ──────────────────────────────────────────────────────────
# Load once at startup. Estimators can edit poc/data/accessories_rules.json
# and the changes take effect on next app restart (or redeploy on Render).

ACCESSORIES_PATH = os.path.join(DATA_DIR, "accessories_rules.json")

def _load_accessories_rules() -> list:
    try:
        data = _load_json(ACCESSORIES_PATH, {})
        rules = data.get("rules", []) if isinstance(data, dict) else []
        # Pre-compile each match term as a word-boundary regex so short
        # abbreviations like "AU" don't match inside unrelated words
        # (e.g. "AUtomatique"). Longer multi-word terms match as-is.
        for r in rules:
            patterns = []
            for m in r.get("match", []):
                m_low = m.lower().strip()
                # Escape special chars, require word boundary on both sides
                # \b works on ASCII word chars; we also accept accented letters
                # by using a lookaround that excludes letters/digits/accents.
                escaped = re.escape(m_low)
                patterns.append(re.compile(
                    r'(?:^|[^\wàâäéèêëîïôöùûüÿç])' + escaped +
                    r'(?:$|[^\wàâäéèêëîïôöùûüÿç])',
                    re.IGNORECASE
                ))
            r["_match_patterns"] = patterns
        print(f"[ACCESSORIES] Loaded {len(rules)} rules from {ACCESSORIES_PATH}")
        return rules
    except Exception as e:
        print(f"[ACCESSORIES] Could not load rules: {e}")
        return []

ACCESSORIES_RULES = _load_accessories_rules()


def get_applicable_accessories(customer_request: str, attachments_text: str = "") -> str:
    """
    Match accessory rules against the customer request (+ any attachments text)
    and return a formatted context block for the LLM prompt.

    Uses word-boundary matching so short terms (e.g. "AU", "ATS") don't trigger
    on substrings inside unrelated words.

    Returns '' if no rules match — keeps the prompt focused.
    """
    if not ACCESSORIES_RULES:
        return ""

    haystack = " " + (customer_request + " " + attachments_text).lower() + " "
    matched = []
    for r in ACCESSORIES_RULES:
        if any(p.search(haystack) for p in r.get("_match_patterns", [])):
            matched.append(r)

    if not matched:
        return ""

    sections = []
    for r in matched:
        lines = [f"▪ {r['label']}  ({r['description']})"]
        for imp in r.get("implies", []):
            lines.append(
                f"    → [{imp['category']}] {imp['item']}   — {imp.get('why','')}"
            )
        if r.get("sizing_impact"):
            lines.append(f"    ⊙ Sizing: {r['sizing_impact']}")
        sections.append("\n".join(lines))

    return (
        "\n=== Accessory knowledge from CETIE estimators (apply when relevant) ===\n"
        "These rules encode the 'when X appears, don't forget Y' knowledge estimators "
        "use daily. They exist ONLY to remind you about accessory items that belong in "
        "the BoM alongside the main components — they are NOT signals about overall "
        "project complexity, price, or labour hours. Use them ADDITIVELY: add the "
        "implied items when applicable and apply any sizing_impact to the enclosure "
        "choice, but DO NOT use them to infer that the whole project is simpler, "
        "cheaper, or smaller than the technical content of the request suggests.\n\n"
        + "\n\n".join(sections)
        + "\n"
    )


def get_relevant_feedback(product_type: str, n: int = 5) -> str:
    """Return recent expert corrections relevant to this product type."""
    feedback = _load_json(FEEDBACK_PATH, [])
    corrections = [f for f in feedback if f.get("rating") == "needs_correction" and f.get("correction_note")]
    # Prioritise same product type, then recency
    def score(f):
        pt = f.get("product_type", "").lower()
        match = 2 if any(w in pt for w in product_type.lower().split()) else 0
        return match
    corrections.sort(key=score, reverse=True)
    recent = corrections[:n]
    if not recent:
        return ""
    lines = "\n".join(
        f"- [{c.get('product_type','?')}] Expert correction: {c['correction_note']}"
        for c in recent
    )
    return f"\nRECENT EXPERT CORRECTIONS on similar configurations (take these into account):\n{lines}\n"

# Build category summaries for the prompt
def get_category_summary():
    cats = {}
    for b in ALL_BLOCKS:
        cat = b["categorie"]
        if cat and cat != "Texte":
            cats.setdefault(cat, []).append(b["designation"])
    # Keep top categories with sample designations
    summary = []
    for cat, desigs in sorted(cats.items()):
        summary.append(f"- {cat}: {desigs[0]}")
    return "\n".join(summary[:80])

BLOCK_CATEGORIES = get_category_summary()

def _auto_cabinet_sizing(configuration: dict, requirements: dict) -> dict:
    """
    Automatically run cabinet sizing after LLM config is generated.
    Extracts inputs from the LLM output — no manual form needed.
    The LLM's spare_reserve_pct is used instead of the hardcoded 20%.
    """
    try:
        bom   = configuration.get("bom_categories", {})
        reqs  = requirements or {}

        # ── Number of motors / pumps ──────────────────────────────────────
        nb_motors = (
            configuration.get("nb_motors")
            or reqs.get("nb_motors")
            or reqs.get("nb_pumps")
            or 1
        )
        try:
            nb_motors = int(nb_motors)
        except Exception:
            nb_motors = 1

        # ── Supply current ────────────────────────────────────────────────
        # Derive from power_kw if not explicit: I ≈ P_kW × 1000 / (400 × √3 × 0.85)
        power_kw = reqs.get("power_kw")
        try:
            supply_current_a = round(float(power_kw) * 1000 / (400 * 1.732 * 0.85)) if power_kw else 63
        except Exception:
            supply_current_a = 63

        # ── Motor feeders from power section ─────────────────────────────
        power_items = bom.get("04_internal_chassis_power", [])
        drive_items = bom.get("04_internal_chassis_automation", [])

        motor_feeders = []
        for it in power_items:
            desig = (it.get("designation") or "").lower()
            if any(kw in desig for kw in ("départ", "disjoncteur moteur", "contacteur moteur", "depart")):
                try:
                    # Extract amperage from designation e.g. "6-10A" → 10
                    import re as _re
                    amps = _re.findall(r"(\d+(?:\.\d+)?)\s*[aA]", desig)
                    amp_val = float(amps[-1]) if amps else 10.0
                    qty = int(it.get("quantity", 1))
                    motor_feeders.extend([{"current_a": amp_val, "type": "direct"}] * qty)
                except Exception:
                    motor_feeders.append({"current_a": 10.0, "type": "direct"})

        if not motor_feeders:
            motor_feeders = [{"current_a": supply_current_a / max(nb_motors, 1), "type": "direct"}] * nb_motors

        # ── VFDs ─────────────────────────────────────────────────────────
        drives = []
        for it in drive_items:
            desig = (it.get("designation") or "").lower()
            if any(kw in desig for kw in ("atv", "variateur", "vfd", "ats")):
                try:
                    import re as _re
                    kws = _re.findall(r"(\d+(?:\.\d+)?)\s*kw", desig)
                    kw_val = float(kws[0]) if kws else 4.0
                    qty = int(it.get("quantity", 1))
                    drives.extend([{"power_kw": kw_val, "has_filter": False}] * qty)
                except Exception:
                    drives.append({"power_kw": 4.0, "has_filter": False})

        # ── PLC type ─────────────────────────────────────────────────────
        plc_type = None
        for it in drive_items:
            desig = (it.get("designation") or "").lower()
            for model in ("m221", "m241", "m340", "s7", "wago", "millenium", "millénium"):
                if model in desig:
                    plc_type = model
                    break
            if plc_type:
                break

        # ── I/O count ────────────────────────────────────────────────────
        automation_desc = str(reqs.get("automation") or "")
        import re as _re
        io_matches = _re.findall(r"(\d+)\s*(?:i/o|e/s|entrée|sortie|io)", automation_desc.lower())
        nb_io = int(io_matches[0]) if io_matches else (16 if plc_type else 0)

        # ── Spare reserve — LLM decides ───────────────────────────────────
        spare_pct = configuration.get("spare_reserve_pct", 20)
        try:
            spare_pct = max(10, min(40, float(spare_pct)))
        except Exception:
            spare_pct = 20.0

        # ── Run sizing ────────────────────────────────────────────────────
        result = cs.calculate_cabinet_sizing(
            supply_current_a=supply_current_a,
            motor_feeders=motor_feeders,
            drives=drives,
            plc_type=plc_type,
            nb_io=nb_io,
            nb_extra_modules=0,
        )

        # Override spare reserve with LLM value
        result["spare_reserve_pct_llm"] = spare_pct
        result["spare_reserve_source"]  = "LLM"

        return result

    except Exception as e:
        print(f"[AutoSizing] Error: {e}")
        return {"error": str(e)}


def search_blocks(keywords: list[str], max_results: int = 20) -> list[dict]:
    """Search blocks by keywords in category or designation."""
    results = []
    keywords_lower = [k.lower() for k in keywords if k]
    seen_ids = set()

    for block in ALL_BLOCKS + ALL_ARMOIRES:
        if block["id"] in seen_ids:
            continue
        text = f"{block['categorie']} {block['designation']} {block['label']}".lower()
        score = sum(1 for kw in keywords_lower if kw in text)
        if score > 0:
            results.append((score, block))
            seen_ids.add(block["id"])

    results.sort(key=lambda x: -x[0])
    return [b for _, b in results[:max_results]]

def enrich_blocks_from_quotes(matching_blocks: list, similar_quotes: list) -> list:
    """
    Add blocks used in similar historical quotes to the candidate list.
    This ensures Claude always sees the exact components from past similar projects,
    even if they didn't rank in the keyword search top-N.
    """
    # Build a fast lookup: id → full block dict
    all_blocks_by_id = {b["id"]: b for b in ALL_BLOCKS + ALL_ARMOIRES}

    existing_ids = {b["id"] for b in matching_blocks}
    enriched     = list(matching_blocks)

    for quote in similar_quotes:
        # Real parsed quotes have 'selected_blocks' with actual IDs
        for sb in quote.get("selected_blocks", []):
            bid = sb.get("id")
            if bid and bid not in existing_ids:
                full_block = all_blocks_by_id.get(bid)
                if full_block:
                    enriched.append(full_block)
                    existing_ids.add(bid)

    return enriched


def get_api_key(provided):
    return provided or os.environ.get("ANTHROPIC_API_KEY", "")

def parse_llm_json(raw: str) -> dict:
    """
    Robustly extract and parse a JSON object from an LLM response.
    Handles: markdown code fences, truncated JSON, French decimal commas,
    literal newlines/tabs inside strings, unescaped special characters.
    Uses json-repair as fallback for anything that survives manual cleanup.
    """
    # 1. Strip markdown code fences
    clean = re.sub(r'```(?:json)?\s*', '', raw).strip().rstrip('`').strip()

    # 2. Find the outermost { ... }
    m = re.search(r'\{.*\}', clean, re.DOTALL)
    if not m:
        return {}
    json_str = m.group()

    # 3. Try clean parse first (fast path)
    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        pass

    # 4. Escape literal control characters inside string values
    result, in_string, i = [], False, 0
    while i < len(json_str):
        c = json_str[i]
        if c == '\\' and in_string:
            result.append(c)
            i += 1
            if i < len(json_str):
                result.append(json_str[i])
            i += 1
            continue
        if c == '"':
            in_string = not in_string
        if in_string:
            if   c == '\n': result.append('\\n')
            elif c == '\r': result.append('\\r')
            elif c == '\t': result.append('\\t')
            else:           result.append(c)
        else:
            result.append(c)
        i += 1
    fixed = ''.join(result)

    try:
        return json.loads(fixed)
    except json.JSONDecodeError:
        pass

    # 5. Last resort: json-repair handles truncated JSON, missing commas,
    #    French decimal separators, unmatched brackets, etc.
    try:
        repaired = repair_json(fixed, return_objects=True)
        print(f"[DEBUG] json-repair result type: {type(repaired)}")
        if isinstance(repaired, dict) and repaired:
            print(f"[DEBUG] json-repair SUCCESS — keys: {list(repaired.keys())}")
            return repaired
        # repair_json sometimes returns a string — parse it
        if isinstance(repaired, str):
            result = json.loads(repaired)
            if isinstance(result, dict):
                print(f"[DEBUG] json-repair string→dict SUCCESS")
                return result
    except Exception as e:
        print(f"[DEBUG] json-repair failed: {e}")

    print(f"[DEBUG] All parsing failed. Raw (first 400 chars):\n{raw[:400]}")
    return {}

# ---------------------------------------------------------------------------
# Demo mode – rule-based fallback when no API key is available
# ---------------------------------------------------------------------------

def demo_extract_requirements(text):
    t = text.lower()

    # Power
    power = None
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*kw', t)
    if m:
        power = float(m.group(1).replace(',', '.'))

    # Nb pumps
    nb_pumps = None
    m = re.search(r'(\d+)\s*pompe', t)
    if m:
        nb_pumps = int(m.group(1))

    # Nb motors
    nb_motors = None
    m = re.search(r'(\d+)\s*moteur', t)
    if m:
        nb_motors = int(m.group(1))

    # Voltage
    voltage = None
    for v in ["400v", "230v", "24v", "48v"]:
        if v in t:
            voltage = v.upper()
            break
    if not voltage and ("triphasé" in t or "tri" in t):
        voltage = "400V"
    if not voltage and ("monophasé" in t or "mono" in t):
        voltage = "230V"

    # IP
    ip = None
    m = re.search(r'ip\s*(\d{2})', t)
    if m:
        ip = f"IP{m.group(1)}"

    # Automation
    automation = None
    for kw in ["s7-1200", "s7-300", "m221", "m241", "millenium", "millénium", "automate", "plc", "api"]:
        if kw in t:
            automation = kw.upper() if len(kw) <= 6 else kw.capitalize()
            break

    # Communication
    comm = None
    for kw in ["modbus", "profibus", "profinet", "ethernet", "canopen"]:
        if kw in t:
            comm = kw.capitalize()
            break

    # Product type
    product_type = "Armoire de commande"
    if "coffret" in t:
        product_type = "Coffret de commande"
    if "pompe" in t:
        product_type += " pompes"
    elif "moteur" in t:
        product_type += " moteurs"
    elif "tgbt" in t:
        product_type = "TGBT"

    # Keywords for block search
    keywords = []
    keyword_map = {
        "pompe": ["pompe", "démarrage"],
        "moteur": ["moteur", "démarrage"],
        "variateur": ["variateur", "atv"],
        "disjoncteur": ["disjoncteur"],
        "arrêt d'urgence": ["urgence", "arrêt"],
        "automatisme": ["automate"],
        "s4w": ["s4w"],
        "s7-1200": ["siemens", "s7"],
        "modbus": ["modbus", "communication"],
        "ip65": ["ip65"],
        "câblage": ["câblage"],
        "alimentation": ["alimentation", "24v"],
    }
    for trigger, kws in keyword_map.items():
        if trigger in t:
            keywords += kws
    # Always add generic ones
    keywords += [w for w in t.split() if len(w) > 4 and w.isalpha()][:5]

    summary = f"Demande de configuration : {product_type}"
    if power:
        summary += f" {power}kW"
    if nb_pumps:
        summary += f", {nb_pumps} pompe(s)"
    if nb_motors:
        summary += f", {nb_motors} moteur(s)"

    return {
        "product_type": product_type,
        "power_kw": power,
        "nb_pumps": nb_pumps,
        "nb_motors": nb_motors,
        "voltage": voltage,
        "protection_ip": ip,
        "automation": automation,
        "communication": comm,
        "special_features": [],
        "keywords": list(set(keywords)),
        "summary": summary,
    }


def demo_build_configuration(requirements, matching_blocks, armoires):
    # Pick enclosure based on nb_pumps / power
    nb = requirements.get("nb_pumps") or requirements.get("nb_motors") or 1
    power = requirements.get("power_kw") or 0

    # Simple heuristic: bigger enclosure for more pumps / higher power
    enc_candidates = [a for a in armoires if a["id"] != 10]  # skip "Texte"
    if nb >= 3 or power > 15:
        enc = enc_candidates[5] if len(enc_candidates) > 5 else enc_candidates[-1]
    elif nb == 2 or power > 5:
        enc = enc_candidates[3] if len(enc_candidates) > 3 else enc_candidates[0]
    else:
        enc = enc_candidates[1] if len(enc_candidates) > 1 else enc_candidates[0]

    # Pick up to 6 relevant blocks (exclude Texte)
    candidate_blocks = [b for b in matching_blocks if b.get("cout", 0) > 0][:6]

    # Fill with sensible defaults if not enough matches
    default_cats = ["Arrêt d'urgence", "Alimentation 24VDC EASY LEGRAND", "Disjoncteur moteur"]
    if len(candidate_blocks) < 3:
        for b in ALL_BLOCKS:
            if b.get("categorie") in default_cats and b not in candidate_blocks:
                candidate_blocks.append(b)
            if len(candidate_blocks) >= 6:
                break

    blocks_out = []
    total_h = enc.get("heures_cablage", 0)
    total_cost = enc.get("cout", 0)
    justifs = [
        "Protection et commande principale",
        "Alimentation auxiliaire des circuits de contrôle",
        "Protection moteur avec réglage thermique",
        "Sécurité opérateur",
        "Signalisation état de marche",
        "Interface de communication",
    ]
    for i, b in enumerate(candidate_blocks):
        qty = nb if i == 0 else 1
        blocks_out.append({
            "id": b["id"],
            "designation": b["designation"],
            "quantity": qty,
            "heures_cablage": b.get("heures_cablage", 0),
            "cost": round(b.get("cout", 0), 2),
            "justification": justifs[i % len(justifs)],
        })
        total_h += b.get("heures_cablage", 0) * qty
        total_cost += b.get("cout", 0) * qty

    missing = []
    if not requirements.get("voltage"):
        missing.append("Tension d'alimentation non précisée")
    if not requirements.get("power_kw"):
        missing.append("Puissance moteur(s) non précisée")
    if not requirements.get("protection_ip"):
        missing.append("Indice de protection (IP) non précisé")

    return {
        "enclosure": {
            "id": enc["id"],
            "designation": enc["designation"],
            "justification": "Sélectionnée selon le nombre d'équipements et la puissance estimée",
        },
        "blocks": blocks_out,
        "total_hours_cablage": round(total_h, 1),
        "total_hours_prog": 1.0 if requirements.get("automation") else 0,
        "estimated_material_cost": round(total_cost, 2),
        "missing_info": missing,
        "assumptions": [
            f"Tension supposée {requirements.get('voltage') or '400V'} triphasé",
            "Schéma de câblage standard CETIE",
            "Disjoncteur de tête inclus",
        ],
        "expert_notes": "[MODE DÉMO – sans IA] Configuration générée par règles heuristiques. À valider et affiner par un expert CETIE avant envoi.",
    }

# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/analyze", methods=["POST"])
def analyze():
    data = request.json
    customer_request = data.get("request", "").strip()
    api_key = get_api_key(data.get("api_key", "").strip())

    if not customer_request:
        return jsonify({"error": "Empty request"}), 400

    demo_mode = not api_key

    if demo_mode:
        # Rule-based fallback – no API needed
        requirements = demo_extract_requirements(customer_request)
        matching_blocks = search_blocks(requirements["keywords"], max_results=30)
        configuration = demo_build_configuration(requirements, matching_blocks, ALL_ARMOIRES)
        return jsonify({
            "requirements": requirements,
            "matching_blocks": matching_blocks[:15],
            "configuration": configuration,
            "demo_mode": True,
        })

    # --- AI mode ---
    client = anthropic.Anthropic(api_key=api_key)

    # Step 0: Ensure RAG index is ready
    _ensure_rag_index()

    # Step 1: Extract requirements
    extraction_prompt = f"""You are an expert at CETIE, a company that manufactures electrical control panels and automation cabinets for industrial applications (pumps, motors, HVAC, water treatment, etc.).

A customer sent this request:
<request>
{customer_request}
</request>

Extract key technical requirements. Respond ONLY with a JSON object like:
{{
  "product_type": "brief product type (e.g. 'armoire de commande pompes', 'coffret automatisme')",
  "power_kw": null or number,
  "nb_pumps": null or number,
  "nb_motors": null or number,
  "voltage": null or "400V" or "230V" etc,
  "protection_ip": null or "IP65" etc,
  "automation": null or brief description,
  "communication": null or "Modbus" etc,
  "special_features": [],
  "keywords": ["list", "of", "french", "technical", "keywords", "for", "component", "search"],
  "summary": "1 sentence summary of what is needed"
}}"""

    extraction_response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=800,
        messages=[{"role": "user", "content": extraction_prompt}]
    )

    try:
        requirements = parse_llm_json(extraction_response.content[0].text)
    except Exception as e:
        print(f"[DEBUG] Extraction parse error: {e}")
        requirements = {"keywords": [], "summary": customer_request[:100]}

    # Step 2: Search relevant blocks
    keywords = requirements.get("keywords", [])
    if requirements.get("product_type"):
        keywords += requirements["product_type"].lower().split()
    if requirements.get("automation"):
        keywords += requirements["automation"].lower().split()

    matching_blocks = search_blocks(keywords, max_results=50)

    # Step 2b: RAG – retrieve similar historical quotes + real DEVIS projects
    similar_quotes   = rag.retrieve_similar(customer_request, n_results=5)
    similar_projects = _retrieve_yearly_projects(customer_request, n_per_year=5)

    # Enrich with blocks actually used in similar past projects
    matching_blocks = enrich_blocks_from_quotes(matching_blocks, similar_quotes)

    blocks_text = "\n".join(
        f"[{b['id']}] {b['categorie']} | {b['designation']} | {b['heures_cablage']}h câblage | €{b['cout']:.2f}"
        for b in matching_blocks
    )
    if similar_quotes:
        rag_context = "\n\n".join(
            f"--- Similar quote #{i+1} (similarity: {q['similarity_score']:.0%}) ---\n"
            f"Request: {q.get('customer_request', '')}\n"
            f"Type: {q.get('product_type', '')} | Sector: {q.get('sector', '')}\n"
            f"Solution: {q.get('summary', '')}\n"
            f"Key blocks: {', '.join(q.get('configuration', {}).get('key_blocks', []))}\n"
            f"Wiring hours: {q.get('configuration', {}).get('total_hours_cablage', '?')}h | "
            f"Prog hours: {q.get('configuration', {}).get('total_hours_prog', 0)}h\n"
            f"Notes: {q.get('configuration', {}).get('notes', '')}"
            for i, q in enumerate(similar_quotes)
        )
        rag_section = f"\nSimilar historical quotes from CETIE (use as inspiration, do NOT copy blindly):\n{rag_context}\n"
    else:
        rag_section = ""

    # Step 3: Generate configuration
    config_prompt = f"""You are a CETIE technical expert configuring electrical control panels.

Customer request:
{customer_request}

Extracted requirements:
{json.dumps(requirements, ensure_ascii=False, indent=2)}
{rag_section}
Available blocks/components (ID | Category | Designation | Wiring hours | Cost):
{blocks_text}

Available enclosures:
{chr(10).join(f"[{a['id']}] {a['categorie']} | {a['designation']} | {a['heures_cablage']}h | €{a['cout']:.2f}" for a in ALL_ARMOIRES[:30])}

Use the similar historical quotes as context to make a better-informed configuration decision.
{get_learned_rules()}{get_relevant_feedback(requirements.get('product_type', ''))}
Propose a technical configuration.

CRITICAL JSON RULES — violation will break the parser:
1. Respond with ONLY the JSON object. No text before or after.
2. Every string value must be on ONE single line. NO literal newlines inside strings.
3. Use decimal DOT (3.5) never comma (3,5) for numbers.
4. Keep each justification to ONE concise sentence (max 20 words).
5. Do not use any special characters that need escaping (tabs, backslashes, quotes inside strings).

{{
  "enclosure": {{
    "id": number or null,
    "designation": "exact designation from the list",
    "justification": "one sentence reason"
  }},
  "blocks": [
    {{
      "id": number,
      "designation": "exact designation from the list",
      "quantity": 1,
      "heures_cablage": number,
      "cost": number,
      "justification": "one sentence reason"
    }}
  ],
  "total_hours_cablage": number,
  "total_hours_prog": number,
  "estimated_material_cost": number,
  "missing_info": ["short item"],
  "assumptions": ["short item"],
  "expert_notes": "one or two sentences maximum",
  "clarification_questions": [
    {{"question": "What IP protection rating is required?", "field": "protection_ip", "type": "choice", "options": ["IP54", "IP55", "IP65", "IP66", "IP67"]}},
    {{"question": "Open question example?", "field": "field_name", "type": "text"}}
  ]
}}

Only include clarification_questions if the answer would significantly change the configuration. Omit the field entirely if the request is already complete."""

    config_response = client.messages.create(
        model="claude-opus-4-7",      # Opus for the quality-critical BoM synthesis step
        max_tokens=8000,
        messages=[{"role": "user", "content": config_prompt}]
    )

    try:
        raw2 = config_response.content[0].text
        configuration = parse_llm_json(raw2)
        if "bom_categories" in configuration:
            configuration["bom_categories"] = cm.match_bom_categories(
                configuration["bom_categories"]
            )
    except Exception as e:
        print(f"[DEBUG] Config parse error: {e}")
        configuration = {"error": f"Could not parse configuration: {e}"}

    return jsonify({
        "requirements":    requirements,
        "matching_blocks": matching_blocks[:15],
        "similar_quotes":  similar_quotes,
        "similar_projects": [
            {
                "id": p.get("id"), "client": p.get("client"),
                "description": p.get("description"),
                "product_type": p.get("product_type"),
                "similarity_score": p.get("similarity_score"),
                "year": p.get("_year") or p.get("year", ""),
                "hours_fabrication": p.get("configuration", {}).get("hours_fabrication", 0),
                "hours_programmation": p.get("configuration", {}).get("hours_programmation", 0),
                "base_price": p.get("configuration", {}).get("base_price", 0),
                "nb_components": p.get("configuration", {}).get("nb_components", 0),
            }
            for p in similar_projects
        ],
        "configuration": configuration,
        "demo_mode": False,
    })

# ---------------------------------------------------------------------------
# Streaming route – SSE progress events + final result
# ---------------------------------------------------------------------------

@app.route("/api/analyze/stream", methods=["POST"])
def analyze_stream():
    data             = request.json or {}
    customer_request = data.get("request", "").strip()
    attachments_text = data.get("attachments_text", "").strip()
    api_key          = get_api_key("")   # always from .env in streaming mode

    def generate():
        def evt(obj):
            return f"data: {json.dumps(obj, ensure_ascii=False)}\n\n"

        if not customer_request:
            yield evt({"type": "error", "msg": "Empty request"})
            return

        if not api_key:
            # Fall back to demo mode instantly
            requirements   = demo_extract_requirements(customer_request)
            matching_blocks = search_blocks(requirements["keywords"], max_results=30)
            configuration  = demo_build_configuration(requirements, matching_blocks, ALL_ARMOIRES)
            yield evt({"type": "result", "data": {
                "requirements": requirements,
                "matching_blocks": matching_blocks[:15],
                "similar_quotes": [],
                "configuration": configuration,
                "demo_mode": True,
            }})
            return

        client = anthropic.Anthropic(api_key=api_key)

        # ── Step 1: Extract requirements ──────────────────────────────────────
        yield evt({"type": "step_start", "step": 1})

        attachments_section = ""
        if attachments_text:
            attachments_section = f"\n\nThe customer also provided the following attached documents. Extract any technical data from them:\n<attachments>\n{attachments_text}\n</attachments>\n"

        extraction_prompt = f"""You are an expert at CETIE, a company that manufactures electrical control panels and automation cabinets for industrial applications (pumps, motors, HVAC, water treatment, etc.).

A customer sent this request:
<request>
{customer_request}
</request>{attachments_section}

Extract key technical requirements. Respond ONLY with a JSON object like:
{{
  "product_type": "brief product type",
  "power_kw": null or number,
  "nb_pumps": null or number,
  "nb_motors": null or number,
  "voltage": null or "400V" or "230V" etc,
  "protection_ip": null or "IP65" etc,
  "automation": null or brief description,
  "communication": null or "Modbus" etc,
  "special_features": [],
  "keywords": ["french", "technical", "keywords"],
  "summary": "1 sentence summary of what is needed"
}}"""

        extraction_response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=800,
            messages=[{"role": "user", "content": extraction_prompt}]
        )
        try:
            requirements = parse_llm_json(extraction_response.content[0].text)
        except Exception as e:
            requirements = {"keywords": [], "summary": customer_request[:100]}

        yield evt({"type": "step_done", "step": 1,
                   "detail": requirements.get("summary", customer_request[:80])})

        # ── Step 2: RAG – retrieve similar quotes & real DEVIS projects ──────
        yield evt({"type": "step_start", "step": 2})

        _ensure_rag_index()
        similar_quotes   = rag.retrieve_similar(customer_request, n_results=5)
        similar_projects = _retrieve_yearly_projects(customer_request, n_per_year=5)

        keywords = requirements.get("keywords", [])
        if requirements.get("product_type"):
            keywords += requirements["product_type"].lower().split()
        if requirements.get("automation"):
            keywords += requirements["automation"].lower().split()
        matching_blocks = search_blocks(keywords, max_results=50)

        # Enrich with blocks actually used in similar past projects
        matching_blocks = enrich_blocks_from_quotes(matching_blocks, similar_quotes)

        nb_refs = len(similar_quotes) + len(similar_projects)
        rag_detail = (
            f"Found {len(similar_quotes)} similar quote(s) + {len(similar_projects)} real DEVIS project(s)"
            if nb_refs else "No similar projects found – using component library only"
        )
        yield evt({"type": "step_done", "step": 2, "detail": rag_detail})

        # ── Step 3: Generate configuration ───────────────────────────────────
        yield evt({"type": "step_start", "step": 3})

        blocks_text = "\n".join(
            f"[{b['id']}] {b['categorie']} | {b['designation']} | {b['heures_cablage']}h | €{b['cout']:.2f}"
            for b in matching_blocks
        )

        # Build RAG context from curated quotes
        if similar_quotes:
            rag_context = "\n\n".join(
                f"--- Similar quote #{i+1} (similarity: {q['similarity_score']:.0%}) ---\n"
                f"Request: {q.get('customer_request','')}\n"
                f"Type: {q.get('product_type','')} | Sector: {q.get('sector','')}\n"
                f"Key blocks: {', '.join(q.get('configuration',{}).get('key_blocks',[]))}\n"
                f"Hours: {q.get('configuration',{}).get('total_hours_cablage','?')}h wiring / "
                f"{q.get('configuration',{}).get('total_hours_prog',0)}h prog"
                for i, q in enumerate(similar_quotes)
            )
            rag_section = f"\n=== Similar historical CETIE quotes (inspiration) ===\n{rag_context}\n"
        else:
            rag_section = ""

        # Build context from real DEVIS projects (yearly data)
        # Pass the FULL BoM (up to 15 items per category) with unit prices so the LLM
        # has concrete components to reference, not just category labels.
        if similar_projects:
            proj_context_parts = []
            for i, proj in enumerate(similar_projects[:3]):
                conf = proj.get("configuration", {})
                cats = conf.get("by_category", {})
                bom_lines = []
                for cat_code, items in cats.items():
                    if items:
                        cat_label = items[0].get("category_label", cat_code)
                        # Up to 15 items per category, with quantity + designation + unit price
                        item_descs = []
                        for it in items[:15]:
                            qty  = it.get("quantity", 1)
                            des  = it.get("designation", "")
                            up   = it.get("unit_price", 0)
                            item_descs.append(f"    - {qty}x {des} @ {up:.2f}€")
                        bom_lines.append(f"  [{cat_code}] {cat_label}:")
                        bom_lines.extend(item_descs)
                        if len(items) > 15:
                            bom_lines.append(f"    ... ({len(items)-15} more)")
                bom_str = "\n".join(bom_lines) if bom_lines else "  (no items)"
                proj_year = proj.get("_year") or proj.get("year", "")
                proj_id   = proj.get("id", "")[:30]  # folder name truncated for reference
                proj_context_parts.append(
                    f"--- Real DEVIS #{i+1} [{proj_year}] ref={proj_id} (similarity: {proj['similarity_score']:.0%}) ---\n"
                    f"Client: {proj.get('client','')} | Description: {proj.get('description','')}\n"
                    f"Fabrication: {conf.get('hours_fabrication',0)}h | Prog: {conf.get('hours_programmation',0)}h | "
                    f"Matière: {conf.get('cost_material',0):.0f}€ | Prix devis: {conf.get('base_price',0):.0f}€\n"
                    f"BoM (full, use these exact designations when applicable):\n{bom_str}"
                )
            yearly_section = (
                "\n=== Real CETIE DEVIS projects — primary reference for BoM structure and component names ===\n"
                + "\n\n".join(proj_context_parts)
                + "\n"
            )
        else:
            yearly_section = ""

        # Build a hint showing which categories similar DEVIS projects populated
        cat_hints = []
        if similar_projects:
            seen_cats = set()
            for proj in similar_projects[:3]:
                for cat, items in proj.get("configuration", {}).get("by_category", {}).items():
                    if items and cat not in seen_cats:
                        seen_cats.add(cat)
            # Normalize 04_internal_chassis → sub-categories
            if "04_internal_chassis" in seen_cats:
                seen_cats.discard("04_internal_chassis")
                seen_cats.add("04_internal_chassis_power")
            if seen_cats:
                cat_hints = sorted(seen_cats)

        cat_hint_str = ""
        if cat_hints:
            cat_hint_str = (
                f"\nIMPORTANT — Similar DEVIS projects used these BoM categories: "
                f"{', '.join(cat_hints)}. "
                f"Your output MUST populate all of these categories with real items.\n"
            )

        accessories_section = get_applicable_accessories(customer_request, attachments_text)

        config_prompt = f"""You are a CETIE technical expert configuring electrical control panels (armoires de commande).

Customer request: {customer_request}{attachments_section}
Extracted requirements:
{json.dumps(requirements, ensure_ascii=False, indent=2)}
{yearly_section}{rag_section}{accessories_section}
Available blocks (ID | Category | Designation | Wiring hours | Cost):
{blocks_text}

Available enclosures:
{chr(10).join(f"[{a['id']}] {a['categorie']} | {a['designation']} | {a['heures_cablage']}h | €{a['cout']:.2f}" for a in ALL_ARMOIRES[:30])}

{get_learned_rules()}{get_relevant_feedback(requirements.get('product_type', ''))}
{cat_hint_str}
GROUNDING — two general rules (apply to every item you output):
1. Component designations must come from either (a) the CETIE catalogue blocks above, or (b) the BoM of a retrieved DEVIS above. Do not invent or paraphrase designations.
2. Every item in bom_categories must include a "source" field: "devis:<ref>" citing the DEVIS it was taken from, or "catalogue" if it comes from the block list. Items without a verifiable source will be rejected downstream.

MANDATORY BoM category rules (always apply):
- 01_cabinet_enclosure: ALWAYS populate — every armoire needs an enclosure.
- 04_internal_chassis_power: ALWAYS populate for motor/pump projects — main switch, protection, busbars.
- 06_door_controls: ALWAYS populate — push buttons, pilot lights, selector switches on the door.
- 11_labor: ALWAYS populate — wiring hours and programming hours as separate line items.
- 02_equipment_on_side: populate if side-panel breakers, door interlocks, or socket outlets are needed.
- 04_internal_chassis_control: populate if relays, timers, or control circuit components are present.
- 04_internal_chassis_automation: populate if a PLC, VFD, or communication module is included.
- Leave a category as [] ONLY if it genuinely does not apply to this project.

Propose a complete technical configuration using CETIE's 12-category BoM structure.

CRITICAL JSON RULES — violation will break the parser:
1. Respond with ONLY the JSON object. No text before or after.
2. Every string value must be on ONE single line. NO literal newlines inside strings.
3. Use decimal DOT (3.5) never comma (3,5) for numbers.
4. Keep each justification to ONE concise sentence (max 20 words).
5. Do not use any special characters that need escaping (tabs, backslashes, quotes inside strings).

{{
  "enclosure": {{"id": number or null, "designation": "exact designation", "justification": "one sentence reason"}},
  "blocks": [{{"id": number, "designation": "exact designation", "quantity": 1, "heures_cablage": number, "cost": number, "justification": "one sentence reason", "bom_category": "01_cabinet_enclosure"}}],
  "bom_categories": {{
    "01_cabinet_enclosure":           [{{"designation": "EXACT designation from DEVIS or catalogue", "quantity": 1, "unit_price": 0, "source": "devis:DEVIS2603141"}}],
    "02_equipment_on_side":           [],
    "04_internal_chassis_power":      [],
    "04_internal_chassis_control":    [],
    "04_internal_chassis_automation": [],
    "05_equipment_on_top":            [],
    "06_door_controls":               [],
    "07_supplied_separately":         [],
    "09_commissioning":               [],
    "10_packaging":                   [],
    "11_labor":                       [{{"designation": "Main d'œuvre câblage", "quantity": 1, "hours": 0, "hourly_rate": 65}}, {{"designation": "Main d'œuvre programmation", "quantity": 1, "hours": 0, "hourly_rate": 75}}],
    "12_options":                     []
  }},
  "total_hours_cablage": number,
  "total_hours_prog": number,
  "estimated_material_cost": number,
  "estimated_price": number,
  "spare_reserve_pct": number,
  "missing_info": ["short item"],
  "assumptions": ["short item"],
  "expert_notes": "one or two sentences maximum",
  "clarification_questions": [
    {{"question": "What is the required IP protection rating?", "field": "protection_ip", "type": "choice", "options": ["IP54", "IP55", "IP65", "IP66", "IP67"]}},
    {{"question": "Short open-ended question?", "field": "field_name", "type": "text"}}
  ]
}}

Fill bom_categories with actual items from the blocks list above, organized by CETIE's 12 categories. Mirror the category structure of the similar DEVIS projects above. Only include clarification_questions if the answer would significantly change the configuration.

For spare_reserve_pct: choose the recommended free space to leave inside the cabinet chassis.
- Simple project (1-2 motors, no automation): 15
- Standard project (2-4 motors, basic PLC): 20
- Complex project (4+ motors, VFDs, advanced automation, telemetry): 25-30"""

        config_response = client.messages.create(
            model="claude-opus-4-7",      # Opus for the quality-critical BoM synthesis step
            max_tokens=8000,
            messages=[{"role": "user", "content": config_prompt}]
        )
        try:
            raw_config = config_response.content[0].text
            print(f"[DEBUG] Config response length: {len(raw_config)} chars, stop_reason: {config_response.stop_reason}")
            configuration = parse_llm_json(raw_config)
            bom = configuration.get("bom_categories", {})
            bom_items = sum(len(v) for v in bom.values() if isinstance(v, list))
            print(f"[DEBUG] bom_categories keys: {list(bom.keys())}, total items: {bom_items}")
            if bom:
                configuration["bom_categories"] = cm.match_bom_categories(bom)
            # Auto-run cabinet sizing using LLM config as input
            configuration["cabinet_sizing"] = _auto_cabinet_sizing(configuration, requirements)
        except Exception as e:
            print(f"[DEBUG] Config parse error: {e}")
            configuration = {"error": str(e)}

        nb_blocks = len(configuration.get("blocks", []))
        yield evt({"type": "step_done", "step": 3,
                   "detail": f"{nb_blocks} components selected · "
                             f"{configuration.get('total_hours_cablage', 0)}h wiring · "
                             f"€{configuration.get('estimated_material_cost', 0):.0f} material"})

        # ── Final result ──────────────────────────────────────────────────────
        yield evt({"type": "result", "data": {
            "requirements":    requirements,
            "matching_blocks": matching_blocks[:15],
            "similar_quotes":  similar_quotes,
            "similar_projects": [
                {
                    "id":          p.get("id"),
                    "client":      p.get("client"),
                    "description": p.get("description"),
                    "product_type": p.get("product_type"),
                    "similarity_score": p.get("similarity_score"),
                    "year":        p.get("_year") or p.get("year", ""),
                    "hours_fabrication": p.get("configuration", {}).get("hours_fabrication", 0),
                    "hours_programmation": p.get("configuration", {}).get("hours_programmation", 0),
                    "base_price": p.get("configuration", {}).get("base_price", 0),
                    "nb_components": p.get("configuration", {}).get("nb_components", 0),
                }
                for p in similar_projects
            ],
            "configuration":   configuration,
            "demo_mode":       False,
        }})

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ---------------------------------------------------------------------------
# Feedback & learning endpoints
# ---------------------------------------------------------------------------

@app.route("/api/feedback", methods=["POST"])
def save_feedback():
    """Save expert feedback on a generated configuration."""
    data = request.json or {}
    feedback = _load_json(FEEDBACK_PATH, [])

    entry = {
        "id":              str(uuid.uuid4())[:8],
        "timestamp":       datetime.now(timezone.utc).isoformat(),
        "user_id":         data.get("user_id", "anonymous"),    # multi-user attribution
        "product_type":    data.get("product_type", ""),
        "request_summary": data.get("request_summary", ""),
        "rating":          data.get("rating", "good"),          # "good" | "needs_correction"
        "correction_note": data.get("correction_note", ""),
        "save_as_rule":    data.get("save_as_rule", False),
        "rule_scope":      data.get("rule_scope", "general"),   # "general" | product type keyword
    }
    feedback.append(entry)
    _save_json(FEEDBACK_PATH, feedback)

    # If expert marked it as a permanent rule, add to learned_rules.json
    if entry["save_as_rule"] and entry["correction_note"]:
        rules = _load_json(RULES_PATH, [])
        rules.append({
            "id":        entry["id"],
            "timestamp": entry["timestamp"],
            "scope":     entry["rule_scope"],
            "rule":      entry["correction_note"],
            "active":    True,
            "source":    "expert_feedback",
        })
        _save_json(RULES_PATH, rules)
        return jsonify({"status": "ok", "saved_as_rule": True, "rule_count": len(rules)})

    return jsonify({"status": "ok", "saved_as_rule": False})


@app.route("/api/rules", methods=["GET"])
def list_rules():
    """Return all learned rules (for the rules manager panel)."""
    return jsonify(_load_json(RULES_PATH, []))


@app.route("/api/rules/<rule_id>", methods=["PATCH"])
def update_rule(rule_id):
    """Toggle a rule active/inactive or update its text."""
    rules = _load_json(RULES_PATH, [])
    data  = request.json or {}
    for r in rules:
        if r["id"] == rule_id:
            if "active" in data:
                r["active"] = data["active"]
            if "rule" in data:
                r["rule"] = data["rule"]
            break
    _save_json(RULES_PATH, rules)
    return jsonify({"status": "ok"})


@app.route("/api/rules/<rule_id>", methods=["DELETE"])
def delete_rule(rule_id):
    """Permanently delete a learned rule."""
    rules = _load_json(RULES_PATH, [])
    rules = [r for r in rules if r["id"] != rule_id]
    _save_json(RULES_PATH, rules)
    return jsonify({"status": "ok"})


# ---------------------------------------------------------------------------
# Cabinet sizing simulation
# ---------------------------------------------------------------------------

@app.route("/api/cabinet-sizing", methods=["POST"])
def cabinet_sizing_route():
    """
    Calculate recommended cabinet dimensions from electrical parameters.

    Body (JSON):
      supply_current_a  : float  – incoming supply current in amps
      motor_feeders     : list   – [{current_a, type}] where type = direct|vfd|soft
      drives            : list   – [{power_kw, has_filter}] for VFDs on chassis
      plc_type          : str|null – 'M221'|'M241'|'S7-1212C'|'S7-1214C'|null
      nb_io             : int    – number of I/O points
      nb_extra_modules  : int    – extra DIN-rail modules (optional, default 0)
    """
    data = request.json or {}

    try:
        supply_current_a  = float(data.get("supply_current_a", 25))
        motor_feeders     = data.get("motor_feeders", [])
        drives            = data.get("drives", [])
        plc_type          = data.get("plc_type") or None
        nb_io             = int(data.get("nb_io", 0))
        nb_extra_modules  = int(data.get("nb_extra_modules", 0))

        result = cs.calculate_cabinet_sizing(
            supply_current_a  = supply_current_a,
            motor_feeders     = motor_feeders,
            drives            = drives,
            plc_type          = plc_type,
            nb_io             = nb_io,
            nb_extra_modules  = nb_extra_modules,
        )
        return jsonify({"status": "ok", "sizing": result})

    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 400


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

@app.route("/api/auth/login", methods=["POST"])
def login():
    data     = request.json or {}
    username = data.get("username", "").strip().lower()
    password = data.get("password", "")
    users    = _load_json(USERS_PATH, [])
    user     = next((u for u in users if u.get("username", "").lower() == username), None)
    if not user or not _verify_password(password, user.get("password_hash", "")):
        return jsonify({"error": "Invalid username or password"}), 401
    token = _secrets.token_hex(32)
    _sessions[token] = user
    _save_sessions()   # persist so the session survives a deploy / restart
    return jsonify({"token": token, "user": _safe_user(user)})

@app.route("/api/auth/logout", methods=["POST"])
def logout():
    token = request.headers.get("X-Auth-Token", "")
    _sessions.pop(token, None)
    _save_sessions()
    return jsonify({"status": "ok"})

@app.route("/api/auth/me", methods=["GET"])
def get_me():
    user = _current_user()
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    return jsonify(_safe_user(user))


@app.route("/api/rag-status", methods=["GET"])
def rag_status():
    """Return which years are indexed and quotes index status."""
    discovered = _discover_years()
    return jsonify({
        "quotes_ready":    _rag_ready,
        "active_years":    _active_years,
        "boot_complete":   _boot_complete,
        "boot_status":     _boot_status,
        "discovered_years": discovered,
        "yearly_data_path": os.path.abspath(YEARLY_DATA_DIR),
    })

@app.route("/api/admin/reindex", methods=["POST"])
def admin_reindex():
    """Admin-only: re-run the full RAG boot (parse + index all years)."""
    _, err = _require_admin()
    if err: return err
    if not _boot_complete:
        return jsonify({"status": "already_running", "message": "Indexing already in progress"}), 409
    # Force-rebuild all yearly indices
    data = request.json or {}
    force_year = data.get("year")  # optional: rebuild specific year only
    if force_year:
        try:
            rag.build_yearly_index(year=force_year, force=True)
            if rag.is_yearly_index_ready(force_year):
                global _active_years
                if force_year not in _active_years:
                    _active_years.append(force_year)
                    _active_years.sort()
            return jsonify({"status": "ok", "year": force_year, "active_years": _active_years})
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500
    else:
        _start_boot()
        return jsonify({"status": "started", "message": "Reindexing all years in background"})


# ---------------------------------------------------------------------------
# Users  (admin only)
# ---------------------------------------------------------------------------

@app.route("/api/users", methods=["GET"])
def list_users():
    _, err = _require_admin()
    if err: return err
    users = _load_json(USERS_PATH, [])
    return jsonify([_safe_user(u) for u in users])

@app.route("/api/users", methods=["POST"])
def create_user():
    _, err = _require_admin()
    if err: return err
    data     = request.json or {}
    username = (data.get("username") or "").strip().lower()
    name     = (data.get("name") or "").strip()
    password = (data.get("password") or "").strip()
    if not username or not name or not password:
        return jsonify({"error": "username, name and password are required"}), 400
    users = _load_json(USERS_PATH, [])
    if any(u.get("username", "").lower() == username for u in users):
        return jsonify({"error": "Username already taken"}), 409
    user = {
        "id":            str(uuid.uuid4())[:8],
        "username":      username,
        "name":          name,
        "role":          data.get("role", "engineer"),
        "password_hash": _hash_password(password),
        "created_at":    datetime.now(timezone.utc).isoformat(),
    }
    users.append(user)
    _save_json(USERS_PATH, users)
    return jsonify(_safe_user(user))

@app.route("/api/users/<user_id>", methods=["PATCH"])
def update_user(user_id):
    _, err = _require_admin()
    if err: return err
    users = _load_json(USERS_PATH, [])
    data  = request.json or {}
    for u in users:
        if u["id"] == user_id:
            if "name"     in data: u["name"]          = data["name"]
            if "role"     in data: u["role"]          = data["role"]
            if "password" in data: u["password_hash"] = _hash_password(data["password"])
            break
    _save_json(USERS_PATH, users)
    return jsonify({"status": "ok"})

@app.route("/api/users/<user_id>", methods=["DELETE"])
def delete_user(user_id):
    _, err = _require_admin()
    if err: return err
    if user_id == "admin":
        return jsonify({"error": "Cannot delete the built-in admin account"}), 400
    users = _load_json(USERS_PATH, [])
    users = [u for u in users if u["id"] != user_id]
    _save_json(USERS_PATH, users)
    return jsonify({"status": "ok"})


# ---------------------------------------------------------------------------
# Attachments & legacy session
# ---------------------------------------------------------------------------

@app.route("/api/parse-attachments", methods=["POST"])
def parse_attachments():
    api_key  = get_api_key(request.form.get("api_key", ""))
    uploaded = request.files.getlist("files")
    if not uploaded:
        return jsonify({"error": "No files uploaded", "files": []}), 400
    results = []
    for f in uploaded:
        data   = f.read()
        parsed = fp.parse_file(f.filename, data, api_key=api_key)
        results.append(parsed)
    return jsonify({"files": results})

@app.route("/api/session", methods=["POST"])
def create_session():
    session_id = str(uuid.uuid4())[:12]
    return jsonify({"session_id": session_id, "created_at": datetime.now(timezone.utc).isoformat()})


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

@app.route("/api/export/excel", methods=["POST"])
def export_excel():
    import io
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter

    data   = request.json or {}
    cfg    = data.get("configuration", {})
    reqs   = data.get("requirements",  {})
    lang   = data.get("lang", "en")
    req_text = data.get("request", "")
    bom    = cfg.get("bom_categories", {})

    BOM_CAT_LABELS = {
        '01_cabinet_enclosure':           ('01 — Enclosure / Cabinet',         '01 — Enveloppe / Armoire'),
        '02_equipment_on_side':           ('02 — Side Equipment',               '02 — Équip. sur côté'),
        '04_internal_chassis':            ('04 — Internal Chassis',             '04 — Châssis intérieur'),
        '04_internal_chassis_power':      ('04a — Chassis / Power',             '04a — Châssis / Puissance'),
        '04_internal_chassis_control':    ('04b — Chassis / Control',           '04b — Châssis / Contrôle'),
        '04_internal_chassis_automation': ('04c — Chassis / Automation',        '04c — Châssis / Automatisme'),
        '05_equipment_on_top':            ('05 — Top Equipment',                '05 — Équip. en toiture'),
        '06_door_controls':               ('06 — Door Controls',                '06 — Commandes façade'),
        '06_door_controls_power':         ('06a — Door Controls / Power',       '06a — Commandes / Puissance'),
        '07_supplied_separately':         ('07 — Supplied Separately',          '07 — Fournitures séparées'),
        '08_electrical_note':             ('08 — Electrical Note',              '08 — Note électrique'),
        '09_commissioning':               ('09 — Commissioning',                '09 — Mise en service'),
        '10_packaging':                   ('10 — Packaging & Transport',        '10 — Emballage & transport'),
        '11_labor':                       ('11 — Labour',                       '11 — Main d\'œuvre'),
        '12_options':                     ('12 — Options',                      '12 — Options & variantes'),
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pre-Configuration"

    # ── Styles ──────────────────────────────────────────────────────────────
    DARK   = "0F1225"
    BLUE   = "1E3A6F"
    ACCENT = "4FC3F7"
    LIGHT  = "EEF4FF"
    WHITE  = "FFFFFF"
    GREY   = "F4F6FA"

    def cell_style(ws, row, col, value="", bold=False, italic=False,
                   font_color=None, bg=None, align="left", size=11, wrap=False, border=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, italic=italic, color=font_color or "1a1a2e", size=size)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if border:
            thin = Side(style='thin', color='D0D8E8')
            c.border = Border(bottom=thin)
        return c

    # Column widths
    ws.column_dimensions['A'].width = 8   # Cat badge
    ws.column_dimensions['B'].width = 55  # Designation
    ws.column_dimensions['C'].width = 8   # Qty
    ws.column_dimensions['D'].width = 14  # Unit price
    ws.column_dimensions['E'].width = 14  # Total
    ws.column_dimensions['F'].width = 14  # Match

    row = 1

    # ── Header block ────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:F{row}')
    c = ws.cell(row=row, column=1, value="CETIE — Technical Pre-Configuration Report")
    c.font  = Font(bold=True, size=16, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 34
    for col in range(2, 7):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=DARK)
    row += 1

    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="AI-Assisted — Preliminary — Expert validation required").font = Font(italic=True, color=ACCENT, size=10)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=BLUE)
    ws.cell(row=row, column=1).alignment = Alignment(indent=1)
    date_str = datetime.now().strftime("%d/%m/%Y")
    ws.merge_cells(f'E{row}:F{row}')
    c = ws.cell(row=row, column=5, value=f"Date: {date_str}")
    c.font = Font(bold=True, color=WHITE, size=10)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="right", indent=1)
    for col in [2,3,4]:
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=BLUE)
    ws.row_dimensions[row].height = 18
    row += 2

    # ── Customer request ────────────────────────────────────────────────────
    if req_text:
        cell_style(ws, row, 1, "CUSTOMER REQUEST", bold=True, font_color=WHITE, bg=BLUE, size=10)
        ws.merge_cells(f'A{row}:F{row}')
        ws.row_dimensions[row].height = 16
        row += 1
        ws.merge_cells(f'A{row}:F{row+2}')
        c = ws.cell(row=row, column=1, value=req_text[:500])
        c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.font = Font(size=10, italic=True, color="2a2a4a")
        ws.row_dimensions[row].height = 60
        row += 4

    # ── Requirements ────────────────────────────────────────────────────────
    req_fields = [
        ("Product Type", "product_type"), ("Power (kW)", "power_kw"),
        ("Pumps", "nb_pumps"), ("Motors", "nb_motors"),
        ("Voltage", "voltage"), ("IP Protection", "protection_ip"),
        ("Automation", "automation"), ("Communication", "communication"),
    ]
    req_vals = [(lbl, reqs.get(k)) for lbl, k in req_fields if reqs.get(k)]
    if req_vals:
        cell_style(ws, row, 1, "TECHNICAL REQUIREMENTS", bold=True, font_color=WHITE, bg=BLUE, size=10)
        ws.merge_cells(f'A{row}:F{row}')
        ws.row_dimensions[row].height = 16
        row += 1
        for lbl, val in req_vals:
            cell_style(ws, row, 1, lbl, bold=False, font_color="6a7a94", bg=GREY, size=10)
            ws.merge_cells(f'A{row}:B{row}')
            cell_style(ws, row, 3, str(val), bold=True, size=10, bg=GREY)
            ws.merge_cells(f'C{row}:F{row}')
            ws.row_dimensions[row].height = 16
            row += 1
        row += 1

    # ── BoM table ────────────────────────────────────────────────────────────
    # Merge A+B for the wide designation column, then C=qty D=unit E=total F=match
    ws.merge_cells(f'A{row}:B{row}')
    for col, (txt, w_align) in enumerate([
        ("Category / Designation", "left"), (None, None), ("Qty", "center"),
        ("Unit Price (€)", "right"), ("Total (€)", "right"), ("Match", "center")
    ], 1):
        if txt is None:
            continue
        c = ws.cell(row=row, column=col, value=txt)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal=w_align, vertical="center", indent=1 if col == 1 else 0)
    ws.row_dimensions[row].height = 20
    row += 1

    grand_total = 0.0
    for catKey, (en_lbl, fr_lbl) in BOM_CAT_LABELS.items():
        items = [it for it in (bom.get(catKey) or []) if it and it.get("designation")]
        if not items:
            continue
        cat_label = fr_lbl if lang == "fr" else en_lbl
        # Category row
        ws.merge_cells(f'A{row}:F{row}')
        c = ws.cell(row=row, column=1, value=f"  {cat_label}")
        c.font = Font(bold=True, size=10, color=DARK)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        for it in items:
            price = float(it.get("unit_price") or it.get("cost") or 0)
            qty   = float(it.get("quantity") or 1)
            total = price * qty
            grand_total += total
            name  = it.get("catalogue_designation") or it.get("designation") or ""
            ms    = it.get("match_status", "not_found")
            badge = {"verified": "✓ cat.", "suggested": "~ approx", "not_found": "?", "skipped": "manual"}.get(ms, "")
            badge_color = {"verified": "166534", "suggested": "854d0e", "not_found": "991b1b", "skipped": "5a5a5a"}.get(ms, "555555")

            is_alt = (row % 2 == 0)
            row_bg = "FAFBFF" if is_alt else WHITE

            ws.merge_cells(f'A{row}:B{row}')
            c = ws.cell(row=row, column=1, value=f"    {name}")
            c.font = Font(size=10, color="1a1a2e")
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="left", vertical="center")

            cell_style(ws, row, 3, int(qty) if qty == int(qty) else qty, align="center", size=10, bg=row_bg)
            cell_style(ws, row, 4, round(price, 2) if price else None, align="right", size=10, bg=row_bg)
            ws.cell(row=row, column=4).number_format = '€#,##0.00'
            cell_style(ws, row, 5, round(total, 2) if total else None, align="right", size=10, bg=row_bg)
            ws.cell(row=row, column=5).number_format = '€#,##0.00'
            c6 = ws.cell(row=row, column=6, value=badge)
            c6.font = Font(size=9, color=badge_color)
            c6.fill = PatternFill("solid", fgColor=row_bg)
            c6.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[row].height = 16
            row += 1

    # Grand total row
    ws.merge_cells(f'A{row}:D{row}')
    c = ws.cell(row=row, column=1, value="TOTAL MATERIAL COST (estimated)")
    c.font = Font(bold=True, size=11, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="left", indent=1)
    c5 = ws.cell(row=row, column=5, value=round(grand_total, 2))
    c5.font = Font(bold=True, size=12, color=WHITE)
    c5.fill = PatternFill("solid", fgColor=DARK)
    c5.number_format = '€#,##0.00'
    c5.alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=6).fill = PatternFill("solid", fgColor=DARK)
    ws.row_dimensions[row].height = 22
    row += 2

    # ── Summary ─────────────────────────────────────────────────────────────
    cell_style(ws, row, 1, "SUMMARY", bold=True, font_color=WHITE, bg=BLUE, size=10)
    ws.merge_cells(f'A{row}:F{row}')
    ws.row_dimensions[row].height = 16
    row += 1
    for lbl, val, fmt in [
        ("Wiring Hours",    cfg.get("total_hours_cablage", 0), "0.0"),
        ("Automation Hours",cfg.get("total_hours_prog", 0),    "0.0"),
        ("Material Cost",   cfg.get("estimated_material_cost", 0), '€#,##0.00'),
        ("Estimated Price", cfg.get("estimated_price", 0),     '€#,##0.00'),
    ]:
        cell_style(ws, row, 1, lbl, font_color="6a7a94", bg=GREY, size=10)
        ws.merge_cells(f'A{row}:D{row}')
        c = ws.cell(row=row, column=5, value=val)
        c.font = Font(bold=True, size=11, color=DARK)
        c.fill = PatternFill("solid", fgColor=GREY)
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=6).fill = PatternFill("solid", fgColor=GREY)
        ws.row_dimensions[row].height = 18
        row += 1

    # ── Footer disclaimer ────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    c = ws.cell(row=row, column=1,
                value="⚠ PRELIMINARY — This document must be validated by a qualified CETIE engineer before client submission. Pricing is indicative only.")
    c.font = Font(italic=True, size=9, color="8a2020")
    c.fill = PatternFill("solid", fgColor="FFF0F0")
    c.alignment = Alignment(wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 28

    # ── Stream to client ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"CETIE_config_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# History  (per-user; admin sees all)
# ---------------------------------------------------------------------------

@app.route("/api/history", methods=["GET"])
def list_history():
    user, err = _require_auth()
    if err: return err
    items = _load_json(HISTORY_PATH, [])
    if user.get("role") != "admin":
        items = [h for h in items if h.get("user_id") == user["id"]]
    summary = [
        {
            "id":              h["id"],
            "title":           h.get("title", "Untitled"),
            "saved_at":        h.get("saved_at", ""),
            "user_id":         h.get("user_id", ""),
            "user_name":       h.get("user_name", ""),
            "product_type":    h.get("product_type", ""),
            "estimated_price": h.get("estimated_price", 0),
        }
        for h in reversed(items)
    ]
    return jsonify(summary)

@app.route("/api/history/<item_id>", methods=["GET"])
def get_history_item(item_id):
    user, err = _require_auth()
    if err: return err
    items = _load_json(HISTORY_PATH, [])
    for h in items:
        if h["id"] == item_id:
            if user.get("role") != "admin" and h.get("user_id") != user["id"]:
                return jsonify({"error": "Forbidden"}), 403
            return jsonify(h)
    return jsonify({"error": "not found"}), 404

@app.route("/api/history", methods=["POST"])
def save_history():
    user, err = _require_auth()
    if err: return err
    data  = request.json or {}
    items = _load_json(HISTORY_PATH, [])
    cfg   = data.get("configuration", {})
    reqs  = data.get("requirements",  {})
    entry = {
        "id":              str(uuid.uuid4())[:12],
        "saved_at":        datetime.now(timezone.utc).isoformat(),
        "title":           data.get("title") or reqs.get("summary", "Untitled")[:80],
        "user_id":         user["id"],
        "user_name":       user["name"],
        "product_type":    reqs.get("product_type", ""),
        "estimated_price": cfg.get("estimated_price", 0),
        "request":         data.get("request", ""),
        "requirements":    reqs,
        "configuration":   cfg,
        "similar_projects": data.get("similar_projects", []),
    }
    items.append(entry)
    _save_json(HISTORY_PATH, items)
    return jsonify({"status": "ok", "id": entry["id"], "title": entry["title"]})

@app.route("/api/history/<item_id>", methods=["DELETE"])
def delete_history_item(item_id):
    user, err = _require_auth()
    if err: return err
    items = _load_json(HISTORY_PATH, [])
    # Non-admin can only delete their own
    new_items = []
    for h in items:
        if h["id"] == item_id:
            if user.get("role") != "admin" and h.get("user_id") != user["id"]:
                return jsonify({"error": "Forbidden"}), 403
            continue   # drop this item
        new_items.append(h)
    _save_json(HISTORY_PATH, new_items)
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    app.run(debug=True, port=5050)
