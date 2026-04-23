#!/usr/bin/env python3
"""
sync_catalogues.py — Refresh blocks.json and armoires.json from the BDD sheets
of the most recent DEVIS .xlsm file.

Why:
  • poc/data/blocks.json / armoires.json were hand-snapshotted in March 2026
    (blocks: 2 661 · armoires: 151)
  • The BDD_Blocs and BDD_Armoires sheets in every 2026 DEVIS file carry the
    LIVE Divalto catalogue (blocks: 2 755 · armoires: 207 as of April 2026)
  • Those sheets grow and have price drift — roughly 1 new block per 2 days
  • Always using the most recent DEVIS's BDDs keeps our catalogues in sync
    without manual re-export from Divalto.

Schema preserved exactly:
  blocks.json    [{id, categorie, designation, heures_cablage, heures_prog, cout, label}]
  armoires.json  [{id, categorie, designation, heures_cablage, cout, label}]

Usage:
  # Default — sync from the newest DEVIS across ALL years
  python3 poc/sync_catalogues.py

  # Only look in a specific year
  python3 poc/sync_catalogues.py --year 2026

  # Dry run — show what WOULD change without overwriting the JSONs
  python3 poc/sync_catalogues.py --dry-run

  # Point at a specific .xlsm
  python3 poc/sync_catalogues.py --file "yearly_data/2026/DEVIS2604122.../...xlsm"
"""
from __future__ import annotations

import argparse
import json
import sys
import warnings
from pathlib import Path
from datetime import datetime

warnings.filterwarnings("ignore")

import openpyxl

BASE_DIR   = Path(__file__).parent.parent
YEARLY_DIR = BASE_DIR / "yearly_data"
DATA_DIR   = Path(__file__).parent / "data"
BLOCKS_JSON   = DATA_DIR / "blocks.json"
ARMOIRES_JSON = DATA_DIR / "armoires.json"


# ── DEVIS file discovery ─────────────────────────────────────────────────────

def find_latest_devis_file(year: str | None = None) -> Path | None:
    """
    Return the path to the most recent DEVIS .xlsm file.
    Priority: highest year → highest DEVIS number → file mtime.
    """
    if not YEARLY_DIR.exists():
        return None

    # Which year dirs to scan
    if year:
        year_dirs = [YEARLY_DIR / year]
    else:
        year_dirs = sorted(
            (d for d in YEARLY_DIR.iterdir()
             if d.is_dir() and d.name.isdigit() and len(d.name) == 4),
            reverse=True,   # newest year first
        )

    candidates: list[tuple[int, Path]] = []

    for year_dir in year_dirs:
        # Some years have a nested year/year/ layout
        target = year_dir / year_dir.name if (year_dir / year_dir.name).exists() else year_dir
        for folder in target.iterdir():
            if not folder.is_dir() or folder.name.startswith("_"):
                continue
            for xlsm in folder.glob("*.xlsm"):
                # Extract devis number from filename
                # e.g. "DEVIS2604122 indice 1.xlsm" → 2604122
                num = 0
                for part in xlsm.stem.split():
                    digits = "".join(c for c in part if c.isdigit())
                    if digits.startswith(("22", "23", "24", "25", "26", "27")):
                        try:
                            num = int(digits[:7]); break
                        except ValueError:
                            pass
                candidates.append((num, xlsm))
        if candidates:  # stop at first year that yielded results
            break

    if not candidates:
        return None
    # Highest devis number = most recent
    candidates.sort(key=lambda x: (-x[0], -x[1].stat().st_mtime))
    return candidates[0][1]


# ── BDD sheet parsers ────────────────────────────────────────────────────────

# Column map (0-based) inside BDD_Blocs and BDD_Armoires
# Verified against DEVIS2604122 — header is on row 7:
#   col 2: id  | col 3: categorie  | col 4: DESIGNATION  | col 5: modules
#   col 6: heures  | col 7: heures_prog  | col 8: cout  | col 10: attribut
C_ID    = 2
C_CAT   = 3
C_DESGN = 4
C_MOD   = 5
C_HOURS = 6
C_HPROG = 7
C_COST  = 8
C_LABEL = 10


def _as_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        try:
            return float(str(v).replace(",", "."))
        except ValueError:
            return 0.0


def _as_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _is_valid_data_row(cat: str, desig: str, cost: float, hours: float) -> bool:
    """
    Reject header rows, reserved slots, and #VALUE! rows.
    A valid record must have:
      • a real designation (not the word "DESIGNATION" which is the column header)
      • a real category (not "Catégorie" which is the column header)
      • EITHER a non-zero cost OR non-zero hours (placeholder rows have both at 0)
    """
    d = desig.strip()
    c = cat.strip()
    if d.upper() == "DESIGNATION":   return False   # header row
    if d == "Texte à remplir":       return False   # template placeholder
    if c.lower() == "catégorie":      return False   # header row
    if not d:                         return False
    if cost <= 0 and hours <= 0:      return False   # reserved empty slot
    return True


def parse_bdd_blocs(wb) -> list[dict]:
    """Extract BDD_Blocs → list matching blocks.json schema."""
    if "BDD_Blocs" not in wb.sheetnames:
        return []
    ws = wb["BDD_Blocs"]
    records: list[dict] = []
    seen_ids: set[int] = set()
    for row in ws.iter_rows(values_only=True):
        if len(row) <= C_COST:
            continue
        bid = _as_int(row[C_ID])
        if bid is None or bid in seen_ids:
            continue
        desig = str(row[C_DESGN] or "").strip()
        cat   = str(row[C_CAT]   or "").strip()
        hours = _as_float(row[C_HOURS])
        cost  = _as_float(row[C_COST])
        if not _is_valid_data_row(cat, desig, cost, hours):
            continue
        records.append({
            "id":             bid,
            "categorie":      cat,
            "designation":    desig,
            "heures_cablage": hours,
            "heures_prog":    _as_float(row[C_HPROG]),
            "cout":           cost,
            "label":          cat,   # blocks.json uses categorie as label
        })
        seen_ids.add(bid)
    return records


def parse_bdd_armoires(wb) -> list[dict]:
    """Extract BDD_Armoires → list matching armoires.json schema."""
    if "BDD_Armoires" not in wb.sheetnames:
        return []
    ws = wb["BDD_Armoires"]
    C_ARM_COST = 7   # 'Cout' column for BDD_Armoires (no heures_prog column)
    records: list[dict] = []
    seen_ids: set[int] = set()
    for row in ws.iter_rows(values_only=True):
        if len(row) <= C_ARM_COST:
            continue
        bid = _as_int(row[C_ID])
        if bid is None or bid in seen_ids:
            continue
        desig = str(row[C_DESGN] or "").strip()
        cat   = str(row[C_CAT]   or "").strip()
        hours = _as_float(row[C_HOURS])
        cost  = _as_float(row[C_ARM_COST])
        if not _is_valid_data_row(cat, desig, cost, hours):
            continue
        records.append({
            "id":             bid,
            "categorie":      cat,
            "designation":    desig,
            "heures_cablage": hours,
            "cout":           cost,
            "label":          desig,   # armoires.json uses designation as label
        })
        seen_ids.add(bid)
    return records


# ── Main driver ──────────────────────────────────────────────────────────────

def print_diff_summary(old: list[dict], new: list[dict], name: str) -> None:
    old_ids = {r["id"] for r in old}
    new_ids = {r["id"] for r in new}
    added   = new_ids - old_ids
    removed = old_ids - new_ids
    kept    = new_ids & old_ids

    old_by_id = {r["id"]: r for r in old}
    new_by_id = {r["id"]: r for r in new}
    price_changed = sum(
        1 for i in kept
        if abs(old_by_id[i].get("cout", 0) - new_by_id[i].get("cout", 0)) > 0.01
    )

    print(f"\n  {name}:")
    print(f"    Before : {len(old)} records")
    print(f"    After  : {len(new)} records   (Δ {len(new) - len(old):+d})")
    print(f"    Added  : {len(added)} new ids")
    print(f"    Removed: {len(removed)} (no longer in BDD)")
    print(f"    Price-changed: {price_changed}")
    if added:
        sample = list(added)[:5]
        for sid in sample:
            r = new_by_id[sid]
            print(f"      + id {sid}: {r['designation'][:60]}  €{r['cout']:.2f}")


def main():
    ap = argparse.ArgumentParser(description="Sync blocks + armoires catalogues from the most recent DEVIS BDD sheets")
    ap.add_argument("--year",    type=str, default=None,
                    help="Year to scan for the newest DEVIS (default: newest year available)")
    ap.add_argument("--file",    type=str, default=None,
                    help="Explicit .xlsm file to extract from (overrides --year)")
    ap.add_argument("--dry-run", action="store_true",
                    help="Show diff but don't overwrite the JSONs")
    args = ap.parse_args()

    # Resolve source file
    if args.file:
        source = Path(args.file).resolve()
        if not source.exists():
            print(f"[ERROR] File not found: {source}")
            sys.exit(1)
    else:
        source = find_latest_devis_file(year=args.year)
        if source is None:
            print(f"[ERROR] No DEVIS .xlsm found in {YEARLY_DIR}"
                  + (f" for year {args.year}" if args.year else ""))
            sys.exit(1)

    print("=" * 70)
    print(" CETIE Catalogue Sync")
    print(f" Source DEVIS: {source.name}")
    print(f" Source path : {source.parent.name}")
    print("=" * 70)

    # Open workbook — read_only for speed
    try:
        wb = openpyxl.load_workbook(str(source), data_only=True, read_only=True)
    except Exception as e:
        # Some files choke openpyxl's autofilter parser — try non-read-only
        print(f"  [WARN] read_only failed ({type(e).__name__}) — retrying in full mode")
        wb = openpyxl.load_workbook(str(source), data_only=True, read_only=False)

    if "BDD_Blocs" not in wb.sheetnames or "BDD_Armoires" not in wb.sheetnames:
        print(f"[ERROR] {source.name} has no BDD sheets. Pick a 2026+ file.")
        print(f"        Available sheets: {wb.sheetnames}")
        sys.exit(1)

    print("\nExtracting…")
    new_blocks   = parse_bdd_blocs(wb)
    new_armoires = parse_bdd_armoires(wb)
    wb.close()

    print(f"  BDD_Blocs    → {len(new_blocks)} block records")
    print(f"  BDD_Armoires → {len(new_armoires)} armoire records")

    # Load existing for diff
    old_blocks   = json.loads(BLOCKS_JSON.read_text())   if BLOCKS_JSON.exists()   else []
    old_armoires = json.loads(ARMOIRES_JSON.read_text()) if ARMOIRES_JSON.exists() else []

    print("\nDiff:")
    print_diff_summary(old_blocks,   new_blocks,   "blocks.json")
    print_diff_summary(old_armoires, new_armoires, "armoires.json")

    if args.dry_run:
        print("\n[DRY RUN] No files written. Remove --dry-run to apply.")
        return

    # Write JSONs
    BLOCKS_JSON.write_text(
        json.dumps(new_blocks, ensure_ascii=False, indent=2)
    )
    ARMOIRES_JSON.write_text(
        json.dumps(new_armoires, ensure_ascii=False, indent=2)
    )

    print(f"\n✓ Wrote {BLOCKS_JSON}  ({len(new_blocks)} records)")
    print(f"✓ Wrote {ARMOIRES_JSON}  ({len(new_armoires)} records)")
    print(f"\nSynced from: {source.name}")
    print(f"Timestamp:   {datetime.now().isoformat(timespec='seconds')}")


if __name__ == "__main__":
    main()
