"""
parse_yearly_data.py
────────────────────
Ingestion pipeline for CETIE yearly project data.
Parses .xlsm (Excel), .msg (Outlook), and .docx files from yearly_data/ folders.
Outputs structured JSON to poc/data/yearly_projects_{year}.json

Usage:
  cd /Users/mac/Projects/Personal/CETIE/projectcetie
  python3 poc/parse_yearly_data.py [year]        # default: 2022
  python3 poc/parse_yearly_data.py 2022 --force  # rebuild even if JSON exists
"""

import os, sys, json, glob, re, traceback
from pathlib import Path
from datetime import datetime

import openpyxl

# ─── Openpyxl tolerance patch ─────────────────────────────────────────────────
# Some 2022 .xlsm files contain autofilter definitions with values that
# openpyxl's strict validator rejects with
# "Value must be either numerical or a string containing a wildcard".
# We patch the descriptor to fall back to raw storage when validation fails,
# since we never actually use those filter values — we just want to read cells.
try:
    from openpyxl.worksheet.filters import CustomFilterValueDescriptor
    _orig_filter_set = CustomFilterValueDescriptor.__set__

    def _tolerant_filter_set(self, instance, value):
        try:
            if isinstance(value, str):
                m = self.pattern.match(value)
                if not m:
                    instance.__dict__[self.name] = value
                    return
                if "*" in value:
                    self.expected_type = str
            from openpyxl.descriptors.base import Typed
            Typed.__set__(self, instance, value)
        except Exception:
            instance.__dict__[self.name] = value

    CustomFilterValueDescriptor.__set__ = _tolerant_filter_set
except Exception:
    pass  # not fatal if openpyxl internals change

try:
    import extract_msg
    HAS_MSG = True
except ImportError:
    HAS_MSG = False
    print("[parse_yearly_data] WARNING: 'extract_msg' not installed — .msg files will be skipped.")
    print("[parse_yearly_data]   Install with: pip install extract-msg")

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

# ─── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent.parent       # projectcetie/
POC_DIR    = Path(__file__).parent              # projectcetie/poc/
DATA_DIR   = POC_DIR / "data"
YEARLY_DIR = BASE_DIR / "yearly_data"

# ─── Chiffrage Specifique — column indices (0-based) ──────────────────────────
# Col 0: always None (padding column)
# Col 1: market flag ('cvc', 'eau', 'dis', …)
# Col 2: '-' for section header  |  numeric ref ID for selected item
# Col 3: section name / item type label
# Col 4: 'o' = optional
# Col 5: reference (often 0)
# Col 6: product designation (actual name)
# Col 7: quantity
# Col 8: unit price (Prix tarif)
# Col 9: discount
# Col 10: total purchase cost
# Col 14: installation / cabling hours
CI_MARKET  = 1
CI_REF     = 2
CI_LABEL   = 3
CI_OPT     = 4
CI_DESGN   = 6
CI_QTY     = 7
CI_PRICE   = 8
CI_TOTAL   = 10
CI_HOURS   = 14

MARKET_FLAGS = {'cvc', 'eau', 'dis', 'CVC', 'EAU', 'DIS'}

# ─── 12-category BoM mapping ───────────────────────────────────────────────────
# Maps sheet section header keywords → CETIE BoM category code
SECTION_CATS = [
    # ── Most specific phrases first ───────────────────────────────────────────
    ('côté armoire',                 '02_equipment_on_side'),
    ('sur côté',                     '02_equipment_on_side'),
    ('commandes en façade',          '06_door_controls_power'),
    ('mise en service',              '09_commissioning'),
    ('fourni séparément',            '07_supplied_separately'),
    ('protection générale',          '04_internal_chassis_power'),
    ('options intérieures',          '04_internal_chassis'),
    ('départs supplémentaires',      '04_internal_chassis_power'),
    ('transformateurs et aliment',   '04_internal_chassis_power'),
    ('mesure de niveau',             '04_internal_chassis_automation'),
    ('mesure de pression',           '04_internal_chassis_automation'),
    ('mesure de débit',              '04_internal_chassis_automation'),
    ('mesure de température',        '04_internal_chassis_automation'),
    ('mesure générale',              '06_door_controls'),
    ('switch ethernet',              '04_internal_chassis_automation'),
    ('convertisseur analogique',     '04_internal_chassis_automation'),
    ('boutonnerie actionneur',       '06_door_controls_power'),
    ('test lampes',                  '06_door_controls'),
    ('ecran s4w',                    '06_door_controls'),
    ('écran s4w',                    '06_door_controls'),
    ('emballage, port',              '10_packaging'),
    ("nombre d'heures sur site",     '09_commissioning'),
    ("type d'intervention",          '09_commissioning'),
    ("type d'intervenant",           '09_commissioning'),
    ('nombre de demi-journée',       '09_commissioning'),
    ("main d'œuvre",                 '11_labor'),
    ("main d'oeuvre",                '11_labor'),
    # ── Single-word / broad matches ───────────────────────────────────────────
    ('enveloppe',            '01_cabinet_enclosure'),
    ('armoire',              '01_cabinet_enclosure'),   # section header 'Armoire'
    ('coffret',              '01_cabinet_enclosure'),   # section header 'Coffret'
    ('socle',                '01_cabinet_enclosure'),
    ('verrine',              '05_equipment_on_top'),
    ('balise',               '05_equipment_on_top'),
    ('gyrophare',            '05_equipment_on_top'),
    ('avertisseur',          '05_equipment_on_top'),
    ('façade',               '06_door_controls'),
    ('porte',                '06_door_controls'),
    ('afficheur',            '06_door_controls'),
    ('voyant',               '06_door_controls'),
    ('boutonnerie',          '06_door_controls'),
    ('actionneurs',          '04_internal_chassis_control'),
    ('démarrage',            '04_internal_chassis_control'),
    ('départ',               '04_internal_chassis_control'),
    ('automate',             '04_internal_chassis_automation'),
    ('télégestion',          '04_internal_chassis_automation'),
    ('télécommande',         '04_internal_chassis_automation'),
    ('communication',        '04_internal_chassis_automation'),
    ('bus',                  '04_internal_chassis_automation'),
    ('ambiance',             '04_internal_chassis'),
    ('protection',           '04_internal_chassis_power'),
    ('alimentation',         '04_internal_chassis_power'),
    ('emballage',            '10_packaging'),
    ('port',                 '10_packaging'),
    ('transport',            '10_packaging'),
    ('prestation',           '09_commissioning'),
    ('intervention',         '09_commissioning'),
    ('département',          '09_commissioning'),
    ('séparément',           '07_supplied_separately'),
    ('prise',                '02_equipment_on_side'),
    ('divers',               '04_internal_chassis'),
    ('option',               '12_options'),
]

CATEGORY_LABELS = {
    '01_cabinet_enclosure':            '01 – Enveloppe / Armoire',
    '02_equipment_on_side':            '02 – Équipements sur côté',
    '04_internal_chassis':             '04 – Châssis intérieur',
    '04_internal_chassis_power':       '04a – Châssis intérieur / Puissance',
    '04_internal_chassis_control':     '04b – Châssis intérieur / Contrôle',
    '04_internal_chassis_automation':  '04c – Châssis intérieur / Automatisme',
    '05_equipment_on_top':             '05 – Équipements en toiture',
    '06_door_controls':                '06 – Commandes en façade',
    '06_door_controls_power':          '06a – Commandes façade / Puissance',
    '07_supplied_separately':          '07 – Fournitures séparées',
    '08_electrical_note':              '08 – Note de calcul électrique',
    '09_commissioning':                '09 – Mise en service sur site',
    '10_packaging':                    '10 – Emballage & transport',
    '11_labor':                        '11 – Main d\'œuvre',
    '12_options':                      '12 – Options & variantes',
}

# ─── Technical keyword detection for tags & architecture text ─────────────────
# (keyword_to_search, normalized_tag)
TECH_KEYWORDS = [
    # Automation brands / PLCs
    ('s7-1200',    'siemens-s7-1200'),  ('s7-300',     'siemens-s7-300'),
    ('s7-',        'siemens'),          ('logo!',       'siemens-logo'),
    ('logo ',      'siemens-logo'),     ('wago',        'wago'),
    ('millenium',  'millenium'),        ('millénium',   'millenium'),
    ('zelio',      'schneider-zelio'),  ('modicon',     'schneider-modicon'),
    ('m340',       'schneider-m340'),   ('beckhoff',    'beckhoff'),
    ('omron',      'omron'),            ('delta',       'delta-plc'),
    # Drives / starters
    ('variateur',  'variateur'),        ('altivar',     'variateur'),
    ('atv3',       'variateur'),        ('atv6',        'variateur'),
    ('atv',        'variateur'),        ('vfd',         'variateur'),
    ('démarreur',  'démarreur'),        ('softstarter', 'démarreur'),
    ('étoile-triangle', 'démarrage-étoile-triangle'),
    # Communication
    ('profinet',   'profinet'),         ('modbus',      'modbus'),
    ('ethernet',   'ethernet'),         ('rs485',       'rs485'),
    ('s4w',        's4w'),              ('gsm',         'gsm'),
    ('gprs',       'gprs'),             ('iot',         'iot'),
    ('lora',       'lora'),
    # Cabinet / IP rating
    ('ip65',       'ip65'),             ('ip54',        'ip54'),
    ('ip55',       'ip55'),             ('ip66',        'ip66'),
    ('ip67',       'ip67'),             ('ip68',        'ip68'),
    ('polyester',  'polyester'),        ('inox',        'inox'),
    ('acier',      'acier'),
    # Applications
    ('step',       'step'),             ('relevage',    'relevage'),
    ('surpresseur','surpresseur'),      ('compresseur', 'compresseur'),
    ('ventilateur','ventilateur'),      ('extracteur',  'extracteur'),
    ('débitmètre', 'débitmètre'),       ('hvac',        'hvac'),
    ('cvc',        'cvc'),
]


def _section_to_category(section_label: str) -> str:
    if not section_label:
        return '04_internal_chassis'
    label_lower = str(section_label).lower()
    for keyword, cat in SECTION_CATS:
        if keyword in label_lower:
            return cat
    return '04_internal_chassis'


# ─── Accueil parser ────────────────────────────────────────────────────────────

def parse_accueil(wb: openpyxl.Workbook) -> dict:
    """Extract project metadata from the Accueil sheet."""
    ws = wb['Accueil']
    meta = {
        'user': None, 'metier': None,
        'ref_affaire': None, 'devis_number': None, 'indice': None,
        'client': None, 'interlocuteur': None,
        'description': None, 'product': None,
        'nb_motors': None,
        'hours_fabrication': 0.0,
        'hours_programmation': 0.0,
        'cost_material': 0.0,
        'base_price': 0.0,
        # Extended fields
        'margin_pct': None,
        'option_price': 0.0,
        'option_hours_fab': 0.0,
        'option_hours_prog': 0.0,
        'divalto_designation': None,
        'has_specifics': False,
        'selling_week': None,
    }

    rows = list(ws.iter_rows(values_only=True))

    def _parse_hours(v) -> float:
        """Parse hours — handles float/int (2022 format) and '20 heures' string (2026+)."""
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        m = re.match(r'(\d+(?:[.,]\d+)?)', str(v).strip())
        return float(m.group(1).replace(',', '.')) if m else 0.0

    # Column headers that appear in the summary header row (row ~27 in 2026 template).
    # Values that look like these should never overwrite real parsed data.
    _HEADER_SENTINELS = {'Client', 'Interlocuteur', 'Affaire', 'Produit',
                         'Commercial', 'Année', 'Semaine envoi', 'N° Devis'}

    for row in rows:
        # Build a lookup of non-None values → original value
        vals = list(row)
        flat = [(i, v) for i, v in enumerate(vals) if v is not None]
        if not flat:
            continue
        flat_v = [v for _, v in flat]
        flat_s = [str(v) for v in flat_v]

        def find_after(keyword):
            """Return value immediately after the cell matching keyword."""
            for j, s in enumerate(flat_s):
                if keyword.lower() in s.lower():
                    if j + 1 < len(flat_v):
                        return flat_v[j + 1]
            return None

        if any('Utilisateur' in s for s in flat_s):
            v = find_after('Utilisateur')
            if v:
                meta['user'] = str(v)
            v = find_after('Métier')
            if v:
                meta['metier'] = str(v)

        if any('Réf Affaire' in s for s in flat_s):
            v = find_after('Réf Affaire')
            if v:
                meta['ref_affaire'] = str(v)

        if any('N° Devis' in s for s in flat_s):
            v = find_after('N° Devis')
            # Only accept real DEVIS numbers — ignore column-header sentinel rows
            if v and not meta['devis_number']:
                vs = str(v).strip()
                if vs not in _HEADER_SENTINELS and (
                        vs.upper().startswith('DEVIS') or re.match(r'^\d{5,}$', vs)):
                    meta['devis_number'] = vs
            # Indice is in "Indice X" cell
            for s in flat_s:
                m = re.search(r'[Ii]ndice\s*(\d+)', s)
                if m:
                    meta['indice'] = int(m.group(1))

        # Offre de base price — checked on every row (2022 and 2026 put it in different rows)
        if not meta['base_price'] and any('Offre de base' in s for s in flat_s):
            v = find_after('Offre de base')
            if v:
                try:
                    meta['base_price'] = float(v)
                except (ValueError, TypeError):
                    pass

        # Client — handle both 'CLIENT' (2022) and 'Client :' (2026) label styles
        if not meta['client'] and any(s.upper().startswith('CLIENT') for s in flat_s):
            v = find_after('Client')
            vs = str(v).strip() if v is not None else ''
            if vs and vs not in _HEADER_SENTINELS:
                meta['client'] = vs

        # Interlocuteur — only accept from rows that have ':' in the label (not header rows)
        if not meta['interlocuteur'] and any(
                'Interlocuteur' in s and ':' in s for s in flat_s):
            v = find_after('Interlocuteur')
            vs = str(v).strip() if v is not None else ''
            if vs and vs not in _HEADER_SENTINELS:
                meta['interlocuteur'] = vs
        # Cost material — appears on 'Interlocuteur' row (2022) or 'Client :' row (2026)
        if not meta['cost_material'] and any('Coût matière' in s for s in flat_s):
            v = find_after('Coût matière')
            if v:
                try:
                    meta['cost_material'] = float(v)
                except (ValueError, TypeError):
                    pass

        if any('Temps fabrication' in s and 'option' not in s.lower() for s in flat_s):
            v = find_after('Temps fabrication')
            if v:
                h = _parse_hours(v)
                if h:
                    meta['hours_fabrication'] = h

        if any('Temps programmation' in s and 'option' not in s.lower() for s in flat_s):
            v = find_after('Temps programmation')
            if v:
                h = _parse_hours(v)
                if h:
                    meta['hours_programmation'] = h

        if any('Nb de moteurs' in s for s in flat_s):
            v = find_after('Nb de moteurs')
            if v:
                try:
                    meta['nb_motors'] = int(v)
                except (ValueError, TypeError):
                    pass

        # MB1 margin %
        if any('MB1' in s for s in flat_s) and meta['margin_pct'] is None:
            v = find_after('MB1')
            if v:
                try:
                    fv = float(v)
                    # Stored as decimal (0.54) or percent (54) — normalise to percent
                    meta['margin_pct'] = round(fv * 100, 1) if fv < 2 else round(fv, 1)
                except (ValueError, TypeError):
                    pass

        # Divalto product designation
        if any('Divalto' in s for s in flat_s) and not meta['divalto_designation']:
            v = find_after('Divalto')
            if v and str(v).strip() not in ('', '-', '0'):
                meta['divalto_designation'] = str(v).strip()

        # Produit field (row ~12)
        if any('Produit' in s and 'Coût' not in s for s in flat_s) and not meta['product']:
            v = find_after('Produit')
            vs = str(v).strip() if v is not None else ''
            if vs and vs not in _HEADER_SENTINELS:
                meta['product'] = vs

        # Has technical specifics?
        if any('Spécificités' in s for s in flat_s):
            v = find_after('Spécificités')
            if v and str(v).strip().upper() in ('OUI', 'YES', '1'):
                meta['has_specifics'] = True

        # Option price
        if any(s == 'Option :' or s == 'Option:' for s in flat_s) and not meta['option_price']:
            v = find_after('Option')
            if v:
                try:
                    meta['option_price'] = float(v)
                except (ValueError, TypeError):
                    pass

        # Option fabrication hours
        if any('Temps fab' in s and 'option' in s.lower() for s in flat_s):
            v = find_after('Temps fab')
            if v:
                h = _parse_hours(v)
                if h:
                    meta['option_hours_fab'] = h

        # Option programmation hours
        if any('Temps prog' in s and 'option' in s.lower() for s in flat_s):
            v = find_after('Temps prog')
            if v:
                h = _parse_hours(v)
                if h:
                    meta['option_hours_prog'] = h

        # Selling week (from summary row e.g. "2026s13")
        if any('Semaine' in s or 'semaine' in s for s in flat_s) and not meta['selling_week']:
            for v in flat_v:
                sv = str(v).strip()
                if re.match(r'\d{4}s\d{1,2}', sv):
                    meta['selling_week'] = sv
                    break

        # Description: row that contains "Temps programmation" also starts with the description
        # (Row 29 in Accueil: ['Armoire 2 pompes 6kW 14A…', 'Temps programmation :', 2, 'H', …])
        if meta['description'] is None and any('Temps programmation' in s and 'option' not in s.lower() for s in flat_s):
            SKIP_PREFIXES = ['Temps', 'Emballage', 'Coût', 'Option', 'Définition',
                             'Interlocuteur', 'Utilisateur', 'CLIENT', 'Réf', 'N°',
                             'Login', 'Société', 'Maintenance', 'Chiffrage', 'Détail']
            for v in flat_v:
                sv = str(v).strip()
                if (len(sv) > 8
                        and not sv[0].isdigit()
                        and not any(sv.startswith(kw) for kw in SKIP_PREFIXES)
                        and sv not in ('H', 'h', '')):
                    meta['description'] = sv
                    break

    return meta


# ─── Chiffrage Specifique parser ───────────────────────────────────────────────

def parse_chiffrage(wb: openpyxl.Workbook) -> dict:
    """
    Parse the Chiffrage sheet (called 'Chiffrage Specifique' in 2022 template,
    'Chiffrage' in 2026+ template).
    Returns a dict of category_code → list of component dicts.
    Also returns totals: total_material_cost, total_hours.
    """
    # Try sheet names in order of preference
    sheet_candidates = ['Chiffrage Specifique', 'Chiffrage']
    ws = None
    for name in sheet_candidates:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        raise KeyError(f"No chiffrage sheet found. Available: {wb.sheetnames}")
    rows = list(ws.iter_rows(values_only=True))

    # ── Auto-detect column positions from header row ───────────────────────────
    # Default to 2022 layout; override if we find a header row
    ci_market = CI_MARKET
    ci_ref    = CI_REF
    ci_label  = CI_LABEL
    ci_opt    = CI_OPT
    ci_desgn  = CI_DESGN
    ci_qty    = CI_QTY
    ci_price  = CI_PRICE
    ci_total  = CI_TOTAL
    ci_hours  = CI_HOURS

    HEADER_KEYWORDS = {
        'désignation': 'ci_desgn',
        'designation': 'ci_desgn',
        'quantité':    'ci_qty',
        'quantite':    'ci_qty',
        'tarif achat': 'ci_price',
        'prix tarif':  'ci_price',
        'achat base':  'ci_total',
        'total':       'ci_total',
    }
    for row in rows:
        vals = [str(v).strip().lower() if v else '' for v in row]
        # Find the header row — use LEFTMOST occurrence of each keyword
        # (2022 sheets repeat headers far to the right for lookup tables)
        matches: dict = {}
        for i, v in enumerate(vals):
            for kw in HEADER_KEYWORDS:
                if kw in v and kw not in matches:
                    matches[kw] = i
        if len(matches) >= 2:
            for kw, col_idx in matches.items():
                var = HEADER_KEYWORDS[kw]
                if var == 'ci_desgn' and col_idx < 20: ci_desgn = col_idx
                elif var == 'ci_qty'  and col_idx < 20: ci_qty   = col_idx
                elif var == 'ci_price' and col_idx < 20: ci_price = col_idx
                elif var == 'ci_total' and col_idx < 20: ci_total = col_idx
            # market, ref, label, opt are always cols 1-4 (never moved)
            break

    categories = {}   # category_code → [component, ...]
    current_section = 'options intérieures'  # default category

    total_material = 0.0
    total_hours    = 0.0

    for row in rows:
        vals = list(row)
        if len(vals) < ci_price + 1:
            continue

        col_market = vals[ci_market]
        col_ref    = vals[ci_ref]
        col_label  = vals[ci_label]
        col_opt    = vals[ci_opt]
        col_desgn  = vals[ci_desgn]
        col_qty    = vals[ci_qty]
        col_price  = vals[ci_price]
        col_total  = vals[ci_total] if len(vals) > ci_total else None
        col_hours  = vals[ci_hours] if len(vals) > ci_hours else None

        # Numeric coercion — 2022 stores values as strings, 2026 as floats
        def _num(v):
            if v is None: return None
            if isinstance(v, (int, float)): return float(v)
            try:    return float(str(v).strip().replace(',', '.'))
            except (ValueError, TypeError): return None

        ref_num   = _num(col_ref)
        qty_num   = _num(col_qty)
        price_num = _num(col_price)
        total_num = _num(col_total)
        hours_num = _num(col_hours)

        # Section header: col_ref == '-' and col_label has text
        if str(col_ref) == '-' and col_label:
            current_section = str(col_label).strip()
            continue

        # Selected component — relaxed criteria:
        # Market flag is optional. Required: numeric ref_id, non-empty designation, qty > 0, price > 0.
        market_ok = col_market is None or col_market == '' or col_market in MARKET_FLAGS
        if (market_ok
                and ref_num is not None
                and ref_num > 0
                and col_desgn
                and str(col_desgn).strip()
                and qty_num is not None
                and qty_num > 0
                and price_num is not None
                and price_num > 0):

            qty   = int(qty_num)
            price = price_num
            total = total_num if total_num is not None else price * qty
            hours = hours_num if hours_num is not None else 0.0
            # Note: 'o' in col_opt = "selectable in template", NOT a commercial option.
            # Commercial options are identified by section header containing 'option'.
            is_template_choice = str(col_opt).strip().lower() == 'o'

            category = _section_to_category(current_section)
            # Override to 12_options only if the section itself is labelled as an option section
            is_option_section = 'option' in current_section.lower()

            component = {
                'ref_id':      int(ref_num),
                'type_label':  str(col_label).strip() if col_label else '',
                'designation': str(col_desgn).strip(),
                'market':      str(col_market),
                'quantity':    qty,
                'unit_price':  round(price, 2),
                'total_price': round(total, 2),
                'hours':       round(hours, 2),
                'optional':    is_option_section,
                'section':     current_section,
                'category':    category,
                'category_label': CATEGORY_LABELS.get(category, category),
            }
            categories.setdefault(category, []).append(component)
            total_material += total
            total_hours    += hours * qty

    return {
        'by_category':     categories,
        'total_material':  round(total_material, 2),
        'total_hours':     round(total_hours, 2),
    }


# ─── E-S (Entrées/Sorties) sheet parser ───────────────────────────────────────

def parse_es_sheet(wb: openpyxl.Workbook) -> dict:
    """
    Parse the E-S (Entrées-Sorties / I/O) sheet.
    Returns counts of digital inputs, digital outputs, analog inputs, analog outputs.
    """
    result = {'digital_in': 0, 'digital_out': 0, 'analog_in': 0, 'analog_out': 0, 'total': 0}
    if 'E-S' not in wb.sheetnames:
        return result

    ws = wb['E-S']
    rows = list(ws.iter_rows(values_only=True))

    # Locate header row containing "Entrées" and "Sorties"
    col_name = col_di = col_do = col_ai = col_ao = None
    data_start = None

    for i, row in enumerate(rows):
        strs = [str(v).lower().strip() if v else '' for v in row]
        if sum(1 for s in strs if 'entrée' in s or 'sortie' in s) >= 2:
            for j, s in enumerate(strs):
                if 'nom' in s and col_name is None:
                    col_name = j
                elif ('entrée' in s or 'entree' in s) and 'ana' not in s and col_di is None:
                    col_di = j
                elif 'sortie' in s and 'ana' not in s and col_do is None:
                    col_do = j
                elif 'ana' in s and ('entrée' in s or 'entree' in s) and col_ai is None:
                    col_ai = j
                elif 'ana' in s and 'sortie' in s and col_ao is None:
                    col_ao = j
            data_start = i + 1
            break

    if data_start is None or col_di is None:
        return result

    def _flag(row, col):
        if col is None or col >= len(row):
            return 0
        v = row[col]
        if isinstance(v, (int, float)) and v:
            return 1
        if str(v).strip().lower() in ('1', 'x', 'oui', 'yes'):
            return 1
        return 0

    for row in rows[data_start:]:
        vals = list(row)
        # Skip completely empty rows
        if not any(v for v in vals if v and str(v).strip()):
            continue
        # Name must be populated
        name_val = vals[col_name] if col_name is not None and col_name < len(vals) else None
        if not name_val or not str(name_val).strip():
            continue
        result['digital_in']  += _flag(vals, col_di)
        result['digital_out'] += _flag(vals, col_do)
        result['analog_in']   += _flag(vals, col_ai)
        result['analog_out']  += _flag(vals, col_ao)

    result['total'] = (result['digital_in'] + result['digital_out']
                       + result['analog_in'] + result['analog_out'])
    return result


# ─── .msg parser ──────────────────────────────────────────────────────────────

def parse_msg_files(folder: Path) -> list[dict]:
    """Parse all .msg files in a folder."""
    if not HAS_MSG:
        return []
    emails = []
    for msg_path in sorted(folder.glob('*.msg')):
        try:
            msg = extract_msg.Message(str(msg_path))
            body = (msg.body or '').strip()
            # Remove excessive whitespace and URL noise
            body = re.sub(r'https?://\S+', '', body)
            body = re.sub(r'\s{3,}', '\n\n', body).strip()

            emails.append({
                'filename': msg_path.name,
                'subject':  (msg.subject or '').strip(),
                'sender':   (msg.sender  or '').strip(),
                'date':     str(msg.date) if msg.date else None,
                'body':     body[:3000],  # cap at 3000 chars
                'attachments': [a.longFilename for a in msg.attachments if a.longFilename],
            })
        except Exception as e:
            print(f"    [WARN] .msg parse failed for {msg_path.name}: {e}")
    return emails


# ─── .docx parser ─────────────────────────────────────────────────────────────

def parse_docx_files(folder: Path) -> list[dict]:
    """Parse all .docx files in a folder for spec text."""
    if not HAS_DOCX:
        return []
    docs = []
    for docx_path in sorted(folder.glob('*.docx')):
        try:
            doc = DocxDocument(str(docx_path))
            text = '\n'.join(p.text for p in doc.paragraphs if p.text.strip())
            docs.append({
                'filename': docx_path.name,
                'text':     text[:4000],
            })
        except Exception as e:
            print(f"    [WARN] .docx parse failed for {docx_path.name}: {e}")
    return docs


# ─── Tag detection ────────────────────────────────────────────────────────────

def _detect_tags(meta: dict, chiffrage: dict, folder_name: str) -> list:
    """Detect relevant tags from all available data sources."""
    tags = set()

    if meta.get('metier'):
        tags.add(meta['metier'].lower().strip())

    # Build a single text from all sources
    all_comps = [
        c.get('designation', '')
        for cat_items in chiffrage.get('by_category', {}).values()
        for c in cat_items
    ]
    all_text = ' '.join([
        meta.get('description') or '',
        meta.get('product') or '',
        meta.get('divalto_designation') or '',
        folder_name,
        ' '.join(all_comps),
    ]).lower()

    for kw, tag in TECH_KEYWORDS:
        if kw.lower() in all_text:
            tags.add(tag)

    # Application type (broader keywords)
    for kw in ['pompe', 'moteur', 'coffret', 'armoire', 'tableau',
               'eau', 'irrigation', 'incendie', 'prestation']:
        if kw in all_text:
            tags.add(kw)

    return sorted(tags)


# ─── Architecture text builder ─────────────────────────────────────────────────

def _build_architecture_text(meta: dict, chiffrage: dict, components: list) -> str:
    """
    Build a rich structured text capturing the TECHNICAL ARCHITECTURE of the project.
    This is used as the primary embedding signal to find truly similar configurations
    (not just similar descriptions).

    Structure: description | parameters | cat01: items | cat04a: items | ... | keywords | hours
    """
    sections = []

    # 1. Product identity
    desc = meta.get('description') or meta.get('product') or ''
    if desc:
        sections.append(desc)

    # 2. Key quantitative parameters
    params = []
    if meta.get('nb_motors'):
        params.append(f"{meta['nb_motors']} moteur{'s' if meta['nb_motors'] > 1 else ''}")
    if meta.get('metier'):
        params.append(meta['metier'])
    if meta.get('divalto_designation'):
        params.append(meta['divalto_designation'])
    if params:
        sections.append(' '.join(params))

    # 3. Architecture by category — the CORE embedding signal
    #    Ordered by technical importance (enclosure → power → control → automation → door → top)
    by_cat = chiffrage.get('by_category', {})
    cat_order = [
        ('01_cabinet_enclosure',           'Enveloppe'),
        ('04_internal_chassis_power',      'Puissance'),
        ('04_internal_chassis_control',    'Contrôle'),
        ('04_internal_chassis_automation', 'Automatisme'),
        ('06_door_controls_power',         'Façade puissance'),
        ('06_door_controls',               'Façade'),
        ('05_equipment_on_top',            'Toiture'),
        ('02_equipment_on_side',           'Sur côté'),
        ('07_supplied_separately',         'Fourni séparément'),
        ('09_commissioning',               'Mise en service'),
    ]
    for cat_key, label in cat_order:
        items = by_cat.get(cat_key, [])
        if not items:
            continue
        desigs = list(dict.fromkeys(
            it['designation'][:50].strip()
            for it in items if it.get('designation')
        ))[:5]
        if desigs:
            sections.append(f"{label}: {' / '.join(desigs)}")

    # 4. Detected technical keywords (brands, protocols, IP, drive type)
    all_text = ' '.join(c.get('designation', '').lower() for c in components) + ' ' + desc.lower()
    detected = []
    for kw, tag in TECH_KEYWORDS:
        if kw.lower() in all_text and tag not in detected:
            detected.append(tag)
    if detected:
        sections.append(' '.join(detected))

    # 5. Hours profile (indicates complexity and automation ratio)
    h_fab  = meta.get('hours_fabrication', 0) or 0
    h_prog = meta.get('hours_programmation', 0) or 0
    if h_fab or h_prog:
        sections.append(f"câblage {h_fab}h automatisme {h_prog}h")

    return ' | '.join(s for s in sections if s)


# ─── Folder processor ─────────────────────────────────────────────────────────

def process_folder(folder: Path):
    """Process a single DEVIS project folder."""
    folder_name = folder.name

    # Find .xlsm file (take first if multiple)
    xlsm_files = sorted(folder.glob('*.xlsm'))
    if not xlsm_files:
        print(f"  [SKIP] No .xlsm file in {folder_name}")
        return None

    xlsm_path = xlsm_files[0]
    print(f"  Parsing {folder_name[:60]}…")

    # Some 2022 .xlsm files have malformed autofilter XML that crashes openpyxl
    # in read_only mode. We probe read_only first (fast); if the probe fails,
    # we load the whole workbook once in full mode and reuse it for every sheet
    # (so we never pay the reload cost twice per file).
    def _load(read_only: bool):
        return openpyxl.load_workbook(str(xlsm_path), data_only=True, read_only=read_only)

    wb = None
    try:
        wb = _load(read_only=True)
        # Probe: force openpyxl to actually parse the problematic sheet once.
        # If it raises the wildcard error, we know we need full mode.
        _probe_sheet = 'Chiffrage Specifique' if 'Chiffrage Specifique' in wb.sheetnames else (
                       'Chiffrage' if 'Chiffrage' in wb.sheetnames else None)
        if _probe_sheet:
            try:
                for _ in wb[_probe_sheet].iter_rows(values_only=True, max_row=50):
                    pass
            except ValueError as e:
                msg = str(e)
                if 'wildcard' in msg or 'numerical' in msg:
                    print(f"    [INFO] Template quirk — loading full mode once")
                    wb.close()
                    wb = _load(read_only=False)
    except Exception as e:
        print(f"    [ERR] Cannot open .xlsm: {e}")
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass
        return None

    # Metadata (Accueil sheet — extended)
    try:
        meta = parse_accueil(wb)
    except Exception as e:
        print(f"    [ERR] Accueil parse: {e}")
        meta = {}

    # Components (Chiffrage sheet)
    try:
        chiffrage = parse_chiffrage(wb)
    except Exception as e:
        print(f"    [ERR] Chiffrage parse: {e}")
        chiffrage = {'by_category': {}, 'total_material': 0.0, 'total_hours': 0.0}

    # I/O signals (E-S sheet)
    try:
        io_data = parse_es_sheet(wb)
    except Exception as e:
        print(f"    [WARN] E-S parse: {e}")
        io_data = {'digital_in': 0, 'digital_out': 0, 'analog_in': 0, 'analog_out': 0, 'total': 0}

    try:
        wb.close()
    except Exception:
        pass
    del wb  # release the full-mode workbook ASAP

    # Emails + docs
    emails = parse_msg_files(folder)
    docs   = parse_docx_files(folder)

    # Build client request string
    client_request_parts = []
    if emails:
        primary = max(emails, key=lambda e: len(e['body']))
        body_clean = re.sub(r'\n+', ' ', primary['body']).strip()
        client_request_parts.append(body_clean[:800])
    if meta.get('description'):
        client_request_parts.append(meta['description'])
    if meta.get('divalto_designation'):
        client_request_parts.append(meta['divalto_designation'])
    client_request = ' | '.join(client_request_parts) or folder_name

    # All selected components (flat list)
    all_components = [c for cat_items in chiffrage['by_category'].values() for c in cat_items]

    # Tags — rich keyword detection across all data sources
    tags = _detect_tags(meta, chiffrage, folder_name)

    # Product type
    product_type = _detect_product_type(meta.get('description') or folder_name)

    # Architecture text — the key embedding improvement
    architecture_text = _build_architecture_text(meta, chiffrage, all_components)

    return {
        'id':                  folder_name,   # unique: full folder name used as stable key
        'folder':              folder_name,
        'year':                folder.parent.name,
        'ref_affaire':         meta.get('ref_affaire'),
        'client':              meta.get('client'),
        'interlocuteur':       meta.get('interlocuteur'),
        'user':                meta.get('user'),
        'metier':              meta.get('metier'),
        # 2022 populates description directly; 2026 only has divalto_designation/product.
        # Fall back through the options so downstream code always sees a non-empty value.
        'description':         (meta.get('description')
                                or meta.get('divalto_designation')
                                or meta.get('product')
                                or ''),
        'product':             meta.get('product'),
        'divalto_designation': meta.get('divalto_designation'),
        'product_type':        product_type,
        'sector':              meta.get('metier', 'EAU'),
        'nb_motors':           meta.get('nb_motors'),
        'has_specifics':       meta.get('has_specifics', False),
        'selling_week':        meta.get('selling_week'),
        'client_request':      client_request,
        'architecture_text':   architecture_text,
        'io': {
            'digital_in':  io_data['digital_in'],
            'digital_out': io_data['digital_out'],
            'analog_in':   io_data['analog_in'],
            'analog_out':  io_data['analog_out'],
            'total':       io_data['total'],
        },
        'emails': emails,
        'docs':   [d['filename'] for d in docs],
        'configuration': {
            'by_category':         chiffrage['by_category'],
            'total_material':      chiffrage['total_material'],
            'total_hours':         chiffrage['total_hours'],
            'hours_fabrication':   meta.get('hours_fabrication', 0),
            'hours_programmation': meta.get('hours_programmation', 0),
            'option_hours_fab':    meta.get('option_hours_fab', 0),
            'option_hours_prog':   meta.get('option_hours_prog', 0),
            'nb_components':       len(all_components),
            'key_components':      [c['designation'] for c in all_components[:8]],
            'base_price':          meta.get('base_price', 0),
            'cost_material':       meta.get('cost_material', 0),
            'option_price':        meta.get('option_price', 0),
            'margin_pct':          meta.get('margin_pct'),
        },
        'tags':       tags,
        'summary':    _build_summary(meta, chiffrage, all_components),
        'parsed_at':  datetime.now().isoformat(),
    }


def _detect_product_type(desc: str) -> str:
    if not desc:
        return 'Armoire de commande'
    d = desc.lower()
    if 'coffret' in d:
        pt = 'Coffret'
    elif 'tableau' in d:
        pt = 'Tableau'
    elif 'armoire' in d:
        pt = 'Armoire'
    else:
        pt = 'Équipement électrique'

    if 'pompe' in d:
        pt += ' commande pompes'
    elif 'step' in d:
        pt += ' STEP'
    elif 'variateur' in d or 'atv' in d:
        pt += ' avec variateur'
    elif 'moteur' in d:
        pt += ' commande moteurs'
    elif 'compresseur' in d:
        pt += ' compresseur'
    elif 'débit' in d:
        pt += ' débitmètre'
    elif 'prestation' in d:
        pt = 'Prestation mise en service'
    return pt


def _build_summary(meta: dict, chiffrage: dict, components: list) -> str:
    parts = []
    if meta.get('client'):
        parts.append(f"Client: {meta['client']}")
    if meta.get('description'):
        parts.append(meta['description'])
    if meta.get('divalto_designation') and meta.get('divalto_designation') != meta.get('description'):
        parts.append(meta['divalto_designation'])
    nb = len(components)
    if nb:
        parts.append(f"{nb} composants")
    if chiffrage.get('total_material'):
        parts.append(f"Matière: {chiffrage['total_material']:.0f}€")
    if meta.get('base_price'):
        parts.append(f"Prix: {meta['base_price']:.0f}€")
    h_fab = meta.get('hours_fabrication', 0) or 0
    h_prg = meta.get('hours_programmation', 0) or 0
    if h_fab or h_prg:
        parts.append(f"{h_fab}h fab + {h_prg}h prog")
    if meta.get('margin_pct'):
        parts.append(f"MB1: {meta['margin_pct']}%")
    return ' | '.join(parts)


# ─── Main ──────────────────────────────────────────────────────────────────────

def run(year: str = '2022', force: bool = False):
    # Support both yearly_data/{year}/{year}/ (old) and yearly_data/{year}/ (new)
    year_dir_nested = YEARLY_DIR / year / year
    year_dir_flat   = YEARLY_DIR / year
    if year_dir_nested.exists():
        year_dir = year_dir_nested
    elif year_dir_flat.exists():
        year_dir = year_dir_flat
    else:
        msg = f"[ERR] Yearly data folder not found: {year_dir_flat} or {year_dir_nested}"
        print(msg)
        # Raise instead of sys.exit so callers (e.g. boot thread) can handle it
        raise FileNotFoundError(msg)

    output_path = DATA_DIR / f"yearly_projects_{year}.json"

    if output_path.exists() and not force:
        print(f"[INFO] {output_path} already exists. Use --force to rebuild.")
        return

    print(f"\n=== CETIE Yearly Data Ingestion — {year} ===")
    print(f"Source: {year_dir}")
    print(f"Output: {output_path}\n")

    folders = sorted(
        p for p in year_dir.iterdir()
        if p.is_dir() and not p.name.startswith('_')
    )
    print(f"Found {len(folders)} project folders\n")

    projects = []
    errors   = []

    for i, folder in enumerate(folders, 1):
        try:
            project = process_folder(folder)
            if project:
                projects.append(project)
        except Exception as e:
            print(f"  [ERR] {folder.name}: {e}")
            errors.append({'folder': folder.name, 'error': str(e)})

    print(f"\n── Results ─────────────────────────────")
    print(f"  Parsed:   {len(projects)} / {len(folders)}")
    print(f"  Errors:   {len(errors)}")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(projects, f, ensure_ascii=False, indent=2)

    print(f"  Saved → {output_path}")

    if errors:
        err_path = DATA_DIR / f"yearly_parse_errors_{year}.json"
        with open(err_path, 'w', encoding='utf-8') as f:
            json.dump(errors, f, ensure_ascii=False, indent=2)
        print(f"  Errors saved → {err_path}")

    # Quick stats
    total_components = sum(p['configuration']['nb_components'] for p in projects)
    clients = {p['client'] for p in projects if p['client']}
    print(f"\n── Stats ───────────────────────────────")
    print(f"  Total components extracted: {total_components}")
    print(f"  Unique clients: {len(clients)}")
    print(f"  Sample clients: {', '.join(sorted(clients)[:8])}")

    return projects


if __name__ == '__main__':
    year  = sys.argv[1] if len(sys.argv) > 1 else '2022'
    force = '--force' in sys.argv
    run(year, force)
