"""
CETIE Excel Quote Parser
========================
Parses a CETIE .xlsm quote file and extracts a structured historical quote
record ready for the RAG knowledge base.

Usage:
    python parse_excel.py                          # parses default file
    python parse_excel.py path/to/DEVIS.xlsm       # parses given file
    python parse_excel.py --rebuild-index          # also rebuilds RAG index
"""

import json
import os
import re
import sys
import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR      = os.path.dirname(__file__)
DATA_DIR      = os.path.join(BASE_DIR, "data")
QUOTES_PATH   = os.path.join(DATA_DIR, "historical_quotes.json")
DEFAULT_EXCEL = os.path.join(
    os.path.dirname(BASE_DIR),
    "DEVIS2601137 indice 1.xlsm"
)

# ── Sheet parsers ──────────────────────────────────────────────────────────────

def parse_accueil(wb: openpyxl.Workbook) -> dict:
    """Extract quote summary from the ACCUEIL sheet."""
    ws   = wb["Accueil"]
    rows = list(ws.iter_rows(min_row=1, max_row=30, values_only=True))

    def cell(r, c):   # 1-based row, 0-based col index
        try:
            return rows[r - 1][c]
        except IndexError:
            return None

    # Parse indice: "Indice 1" → "1"
    raw_indice = cell(13, 4)
    indice = str(raw_indice).replace("Indice", "").strip() if raw_indice else "1"

    return {
        "devis_number":        cell(13, 3),   # col D → 'DEVIS2601137'
        "indice":              indice,         # col E → 'Indice 1' → '1'
        "client":              cell(9,  3),   # col D → 'HYDRELEC'
        "interlocuteur":       cell(10, 3),   # col D → 'M. Jacques FOUILLERON'
        "ref_affaire":         cell(11, 3),   # col D → 'Bisinchi'
        "produit":             cell(12, 3),   # col D → 'Armoire 2 pompes 2kW avec S4W'
        "commercial":          cell(8,  3),   # col D → 'Thierry GUYOT'
        "prix_vente":          cell(8,  8),   # col I → 7614
        "cout_matiere":        cell(9,  8),   # col I → 4210.28
        "mb1_pct":             cell(10, 8),   # col I → 0.447
        "temps_fabrication":   cell(15, 8),   # col I → '22 heures'
        "temps_programmation": cell(16, 8),   # col I → '1 heures'
    }


def parse_selected_blocks(wb: openpyxl.Workbook) -> list[dict]:
    """
    Extract all selected components from CHIFFRAGE.
    A component is 'selected' when:  qty > 0  AND  unit purchase cost > 0
    """
    ws       = wb["Chiffrage"]
    selected = []

    for row in ws.iter_rows(min_row=14, values_only=True):
        block_id = row[2]    # Colonne1  – block ID
        label    = row[3]    # Libellé court
        desig    = row[5]    # Désignation (full)
        qty      = row[6]    # Quantité
        cost     = row[7]    # Tarif achat (unit purchase price)
        hrs_w    = row[16]   # Heures atelier base
        hrs_p    = row[18]   # Heures autom base
        cat      = row[34]   # Category string (01.Enveloppe …)

        is_selected = (
            qty  and isinstance(qty,  (int, float)) and qty  > 0
            and cost and isinstance(cost, (int, float)) and cost > 0
        )
        if not is_selected:
            continue

        display = str(desig).strip() if desig else (str(label).strip() if label else "")

        selected.append({
            "id":               block_id,
            "label":            str(label).strip() if label else "",
            "designation":      display,
            "quantity":         qty,
            "unit_cost":        round(float(cost), 4),
            "heures_cablage":   round(float(hrs_w), 4) if isinstance(hrs_w, (int, float)) else 0.0,
            "heures_prog":      round(float(hrs_p), 4) if isinstance(hrs_p, (int, float)) else 0.0,
            "categorie":        str(cat).strip() if cat else "",
        })

    return selected


# ── Tag inference ──────────────────────────────────────────────────────────────

_PRODUCT_KEYWORDS = [
    "pompe", "moteur", "variateur", "compresseur", "tgbt", "ventilation",
    "automatisme", "télégestion", "teledgestion", "s4w", "sofrel", "s7-1200",
    "m241", "m221", "relevage", "surpresseur", "irrigation", "hydraulique",
    "cta", "climatisation", "portail", "éclairage",
]

_BLOCK_KEYWORDS = [
    "pompe", "variateur", "s4w", "sofrel", "gsm", "4g", "modbus", "profibus",
    "profinet", "siemens", "schneider", "flotteur", "pression", "débit",
    "ip65", "ip54", "ip55", "400v", "230v", "24v", "alternance", "permutation",
    "urgence", "sécurité", "relais", "contacteur", "disjoncteur",
]


def infer_tags(produit: str, blocks: list[dict]) -> list[str]:
    tags = set()
    text = (produit or "").lower()

    for kw in _PRODUCT_KEYWORDS:
        if kw in text:
            tags.add(kw)

    for b in blocks:
        blob = (b.get("designation", "") + " " + b.get("label", "")).lower()
        for kw in _BLOCK_KEYWORDS:
            if kw in blob:
                tags.add(kw)

    # Power
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*kw', text)
    if m:
        tags.add(f"{m.group(1).replace(',', '.')}kW")

    # Pump count
    m2 = re.search(r'(\d+)\s*(?:pompe|ppes?)', text)
    if m2:
        tags.add(f"{m2.group(1)} pompes")

    return sorted(tags)


# ── Hours helper ───────────────────────────────────────────────────────────────

def _parse_hours(val) -> float:
    if isinstance(val, (int, float)):
        return round(float(val), 1)
    if isinstance(val, str):
        m = re.search(r'(\d+(?:[.,]\d+)?)', val)
        if m:
            return round(float(m.group(1).replace(",", ".")), 1)
    return 0.0


# ── Sector detection ───────────────────────────────────────────────────────────

def _detect_sector(tags: list[str], produit: str) -> str:
    t = set(tags)
    p = (produit or "").lower()
    if any(k in t for k in ["ventilation", "cta", "climatisation"]):
        return "CVC / Bâtiment"
    if any(k in t for k in ["hydraulique"]):
        return "Industrie / Machines spéciales"
    if "tgbt" in p:
        return "Industrie générale"
    if any(k in t for k in ["relevage", "surpresseur", "irrigation", "s4w", "sofrel"]):
        return "Eau / Assainissement"
    if any(k in t for k in ["pompe", "moteur"]):
        return "Industrie / Process"
    return "Industrie générale"


# ── Main builder ───────────────────────────────────────────────────────────────

def build_quote_record(header: dict, blocks: list[dict], quote_id: int) -> dict:
    """Combine parsed header + blocks into a historical_quotes.json entry."""

    produit  = str(header.get("produit")      or "").strip()
    client   = str(header.get("client")       or "").strip()
    ref      = str(header.get("ref_affaire")  or "").strip()
    devis_no = str(header.get("devis_number") or "").strip()
    indice   = str(header.get("indice")       or "1").strip()

    hrs_fab  = _parse_hours(header.get("temps_fabrication"))
    hrs_prog = _parse_hours(header.get("temps_programmation"))
    cout_mat = round(float(header.get("cout_matiere") or 0), 2)
    prix_vte = round(float(header.get("prix_vente")   or 0), 2)

    # Enclosure = first selected item from the 01.Enveloppe category
    enclosure = next(
        (b["designation"] for b in blocks
         if "enveloppe" in (b.get("categorie") or "").lower()
         and b.get("unit_cost", 0) > 1),
        produit,
    )

    # Key blocks: exclude enclosure, labour, packaging, near-zero cost text lines
    key_blocks = [
        b["designation"]
        for b in blocks
        if b.get("unit_cost", 0) > 1
        and not any(k in (b.get("categorie") or "").lower()
                    for k in ["enveloppe", "main d", "emballage", "divers"])
    ]

    tags   = infer_tags(produit, blocks)
    sector = _detect_sector(tags, produit)

    # Build a rich text that represents what a customer might request
    customer_request = (
        f"Demande de devis pour : {produit}. "
        f"Client : {client}. Affaire : {ref}. "
        f"Composants principaux : {', '.join(key_blocks[:10])}."
    )

    summary = (
        f"{produit} – {client} ({ref}) | "
        f"{hrs_fab}h câblage / {hrs_prog}h prog | "
        f"Matière : {cout_mat} EUR | Prix vente : {prix_vte} EUR"
    )

    return {
        "id":               quote_id,
        "source_file":      devis_no,
        "indice":           indice,
        "customer_request": customer_request,
        "product_type":     produit,
        "sector":           sector,
        "summary":          summary,
        "tags":             tags,
        "client":           client,
        "ref_affaire":      ref,
        "selected_blocks":  blocks,           # full detail for reference
        "configuration": {
            "enclosure":                enclosure,
            "key_blocks":               key_blocks,
            "total_hours_cablage":      hrs_fab,
            "total_hours_prog":         hrs_prog,
            "estimated_material_cost":  cout_mat,
            "estimated_sale_price":     prix_vte,
            "notes": (
                f"Devis réel CETIE n°{devis_no} indice {indice}. "
                f"Commercial : {header.get('commercial')}."
            ),
        },
    }


# ── Quotes JSON management ─────────────────────────────────────────────────────

def load_quotes() -> list[dict]:
    if os.path.exists(QUOTES_PATH):
        with open(QUOTES_PATH, encoding="utf-8") as f:
            return json.load(f)
    return []


def save_quotes(quotes: list[dict]) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(QUOTES_PATH, "w", encoding="utf-8") as f:
        json.dump(quotes, f, ensure_ascii=False, indent=2)


def upsert_quote(new_quote: dict) -> tuple[list[dict], bool]:
    """Add or replace the quote with the same source_file + indice."""
    quotes  = load_quotes()
    key     = (new_quote["source_file"], new_quote["indice"])
    replaced = False

    for i, q in enumerate(quotes):
        if (q.get("source_file"), q.get("indice")) == key:
            quotes[i]  = new_quote
            replaced   = True
            break

    if not replaced:
        # Assign a new unique id
        existing_ids = {q["id"] for q in quotes}
        new_id = max(existing_ids, default=0) + 1
        new_quote["id"] = new_id
        quotes.insert(0, new_quote)   # real quotes go first

    return quotes, replaced


# ── Entry point ────────────────────────────────────────────────────────────────

def parse_file(excel_path: str, rebuild_index: bool = False) -> dict:
    print(f"[PARSER] Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True,
                                keep_vba=True, data_only=True)

    header = parse_accueil(wb)
    blocks = parse_selected_blocks(wb)
    wb.close()

    print(f"[PARSER] Quote    : {header.get('devis_number')} indice {header.get('indice')}")
    print(f"[PARSER] Product  : {header.get('produit')}")
    print(f"[PARSER] Client   : {header.get('client')} – {header.get('ref_affaire')}")
    print(f"[PARSER] Hours    : {_parse_hours(header.get('temps_fabrication'))}h cab "
          f"/ {_parse_hours(header.get('temps_programmation'))}h prog")
    print(f"[PARSER] Blocks   : {len(blocks)} selected components")

    quote   = build_quote_record(header, blocks, quote_id=0)
    quotes, replaced = upsert_quote(quote)
    save_quotes(quotes)
    action = "updated" if replaced else "added"
    print(f"[PARSER] Quote {action} in {QUOTES_PATH}  ({len(quotes)} total quotes)")

    if rebuild_index:
        print("[PARSER] Rebuilding RAG index …")
        # Load .env
        env_path = os.path.join(BASE_DIR, ".env")
        if os.path.exists(env_path):
            with open(env_path) as ef:
                for line in ef:
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        k, v = line.split("=", 1)
                        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        import rag
        rag.build_index(force=True)

    return quote


if __name__ == "__main__":
    args        = sys.argv[1:]
    rebuild     = "--rebuild-index" in args
    excel_paths = [a for a in args if not a.startswith("--")]
    path        = excel_paths[0] if excel_paths else DEFAULT_EXCEL

    if not os.path.exists(path):
        print(f"[ERROR] File not found: {path}")
        sys.exit(1)

    quote = parse_file(path, rebuild_index=rebuild)

    print("\n[PARSER] Resulting record (summary):")
    print(f"  product_type : {quote['product_type']}")
    print(f"  sector       : {quote['sector']}")
    print(f"  tags         : {quote['tags']}")
    print(f"  key_blocks   : {len(quote['configuration']['key_blocks'])} items")
    print(f"  hours        : {quote['configuration']['total_hours_cablage']}h câblage "
          f"/ {quote['configuration']['total_hours_prog']}h prog")
    print(f"  material     : {quote['configuration']['estimated_material_cost']} EUR")
    print(f"  sale price   : {quote['configuration']['estimated_sale_price']} EUR")
